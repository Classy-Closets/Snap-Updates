
import bpy
from bpy.types import Panel
from bpy.types import Operator
from bpy.props import (StringProperty,
                       BoolProperty,
                       IntProperty,
                       FloatProperty,
                       FloatVectorProperty,
                       BoolVectorProperty,
                       PointerProperty,
                       CollectionProperty,
                       EnumProperty)
import os, sys, subprocess
from snap import sn_types, sn_unit, sn_utils, sn_db, sn_paths, sn_xml
from . import closet_props
from . import closet_utils
import xml.etree.ElementTree as ET

PRICING_PROPERTY_NAMESPACE = "sn_project_pricing"

SKU_NUMBER = ''
PART_LABEL_ID = ''
PART_NAME = ''
QUANTITY = ''
LENGTH = ''
WIDTH = ''
THICKNESS = ''
DESCRIPTION = ''

NAME = ''
TYPE = ''
VALUE = ''

#Counters for edgebanding field
S_COUNT = []
L_COUNT = []
EDGEBANDING = []

R_MATERIAL_PRICES = []
R_HARDWARE_PRICES = []
R_ACCESSORY_PRICES = []
R_WOOD_PANEL_PRICES = []
R_SPECIAL_ORDER_PRICES = []
R_LABOR_PRICES = []
R_MATERIAL_SQUARE_FOOTAGE = []
R_MATERIAL_LINEAR_FOOTAGE = []
R_PROJECT_TOTAL_HARDWARE = []
R_PROJECT_TOTAL_ACCESSORIES = []
R_PROJECT_TOTAL_MATERIAL = []
R_PROJECT_TOTAL_SQUARE_FOOTAGE = []
R_PROJECT_TOTAL_LINEAR_FOOTAGE = []
R_PROJECT_TOTAL_PRICE = []
R_ROOM_PRICING_LIST = []
R_PROJECT_TOTAL_WOOD_PANEL = []
R_PROJECT_TOTAL_LABOR = []

F_MATERIAL_PRICES = []
F_HARDWARE_PRICES = []
F_ACCESSORY_PRICES = []
F_WOOD_PANEL_PRICES = []
F_SPECIAL_ORDER_PRICES = []
F_LABOR_PRICES = []
F_MATERIAL_SQUARE_FOOTAGE = []
F_MATERIAL_LINEAR_FOOTAGE = []
F_PROJECT_TOTAL_HARDWARE = []
F_PROJECT_TOTAL_ACCESSORIES = []
F_PROJECT_TOTAL_MATERIAL = []
F_PROJECT_TOTAL_SQUARE_FOOTAGE = []
F_PROJECT_TOTAL_LINEAR_FOOTAGE = []
F_PROJECT_TOTAL_PRICE = []
F_ROOM_PRICING_LIST = []
F_PROJECT_TOTAL_WOOD_PANEL = []
F_PROJECT_TOTAL_LABOR = []

MATERIAL_PARTS_LIST = []
HARDWARE_PARTS_LIST = []
ACCESSORY_PARTS_LIST = []
WOOD_PANEL_PARTS_LIST = []
SPECIAL_ORDER_PARTS_LIST = []

assembly_types = ['shelf'.upper(), 'topshelf'.upper(), 'cleat'.upper(), 'back'.upper(), 'face'.upper(),
                  'front'.upper(), 'bottom'.upper(), 'side'.upper(), 'door'.upper(), 'drawer'.upper()]
material_types = ['PM', 'VN', 'EB', 'MD', 'WD', 'RE', 'GL', 'SN', 'PL', 'BB']
hardware_types = ['HW']
accessory_types = ['AC', 'WB', 'CM']
special_order_types = ['SO']
labor_types = ['PM']
eb_orientation = ''

def set_job_info(name, var):
    # Job Information
    global JOB_NAME
    global JOB_NUMBER
    global CLIENT_NAME
    global CLIENT_ID
    global CLIENT_ADDRESS
    global CLIENT_CITY
    global CLIENT_STATE
    global CLIENT_ZIP
    global CLIENT_PHONE_1
    global CLIENT_PHONE_2
    global CLIENT_EMAIL
    global CLIENT_NOTES
    global CLIENT_DESIGNER
    global CLIENT_ROOM_COUNT
    global DESIGN_DATE

    if name.text is not None:
        JOB_NAME = name.text
    else:
        JOB_NAME = ''
    if var[0][1].text is not None:
        JOB_NUMBER = var[0][1].text
    else:
        JOB_NUMBER = ''
    if var[1][1].text is not None:
        CLIENT_NAME = var[1][1].text
    else:
        CLIENT_NAME = ''
    if var[2][1].text is not None:
        CLIENT_ID = var[2][1].text
    else:
        CLIENT_ID = ''
    if var[3][1].text is not None:
        CLIENT_ADDRESS = var[3][1].text
    else:
        CLIENT_ADDRESS = ''
    if var[4][1].text is not None:
        CLIENT_CITY = var[4][1].text
    else:
        CLIENT_CITY = ''
    if var[5][1].text is not None:
        CLIENT_STATE = var[5][1].text
    else:
        CLIENT_STATE = ''
    if var[6][1].text is not None:
        CLIENT_ZIP = var[6][1].text
    else:
        CLIENT_ZIP = ''
    if var[7][1].text is not None:
        CLIENT_PHONE_1 = var[7][1].text
    if var[8][1].text is not None:
        CLIENT_PHONE_2 = var[8][1].text
    if var[9][1].text is not None:
        CLIENT_EMAIL = var[9][1].text
    if var[10][1].text is not None:
        CLIENT_NOTES = var[10][1].text
    if var[11][1].text is not None:
        CLIENT_DESIGNER = var[11][1].text
    else:
        CLIENT_DESIGNER = ''
    if var[12][1].text is not None:
        CLIENT_ROOM_COUNT = var[12][1].text
    else:
        CLIENT_ROOM_COUNT = ''
    if var[13][1].text is not None:
        DESIGN_DATE = var[13][1].text
    else:
        DESIGN_DATE = ''


def get_project_xml(self):
    props = bpy.context.window_manager.sn_project
    proj = props.get_project()
    cleaned_name = proj.get_clean_name(proj.name)
    project_dir = bpy.context.preferences.addons['snap'].preferences.project_dir
    selected_project = os.path.join(project_dir, cleaned_name)
    xml_file = os.path.join(selected_project, "snap_job.xml")
    global PROJECT_NAME
    PROJECT_NAME = cleaned_name

    if not os.path.exists(project_dir):
        print("Projects Directory does not exist")
        # os.makedirs(project_dir)
    if not os.path.exists(xml_file):
        print("The 'snap_job.xml' file is not found. Please select desired rooms and perform an export.")   
    else:
        return xml_file


def get_pricing_props():
    """ 
    This is a function to get access to all of the scene properties that are registered in this library
    """
    props = eval("bpy.context.scene." + PRICING_PROPERTY_NAMESPACE)
    return props


def get_square_footage(length_inches, width_inches):
    return round((width_inches * length_inches) / 144, 2)


def get_linear_footage(length_inches):
    return round((length_inches) / 12, 2)


def price_check(sku_num, franchise, retail):
    if franchise > retail:
        print("Price discrepancy within database")
        print("Sku Number: " + str(sku_num) + "  Franchise Price: " + str(franchise) + "  Retail Price: " + str(retail))


def set_column_width(sheet):
    try:
        import openpyxl
    except ModuleNotFoundError:
        python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
        sys.path.append(python_lib_path)

    ws = sheet
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.row == 1:
                cell.font = openpyxl.styles.Font(bold=True)
            if cell.value or cell.value == 0 or cell.value == '0':
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 3
    # Code to freeze header row
    ws.freeze_panes = 'A2'


def display_parts_summary(parts_file):
    os.startfile(parts_file)


def generate_retail_pricing_summary(parts_file):
    try:
        import openpyxl
        import et_xmlfile
        import pandas
    except ModuleNotFoundError:
        python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
        sys.path.append(python_lib_path)

    wb = openpyxl.Workbook()
    pricing_sheet = wb.active
    pricing_sheet.title = "Retail Pricing Summary"
    pricing_sheet.HeaderFooter.oddHeader.left.text = "Client Name: {}\nClient ID: {}".format(CLIENT_NAME, CLIENT_ID)
    pricing_sheet.HeaderFooter.oddHeader.center.text = "Pricing Summary Sheet\nJob Number: {}".format(JOB_NUMBER)
    pricing_sheet.HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
    row_start = 0

    for i in range(len(R_ROOM_PRICING_LIST)):
        pricing_sheet["A" + str(row_start + 1)] = "Room Name"
        pricing_sheet["A" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 1)] = R_ROOM_PRICING_LIST[i][0]
        pricing_sheet["C" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["D" + str(row_start + 1)] = "Discount %"
        pricing_sheet["D" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["E" + str(row_start + 1)] = "Discount Value"
        pricing_sheet["E" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["A" + str(row_start + 2)] = "Special Order Items Count"
        pricing_sheet["A" + str(row_start + 2)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["B" + str(row_start + 2)] = "See Special Order tab for details"
        pricing_sheet["C" + str(row_start + 2)] = str(R_ROOM_PRICING_LIST[i][7])
        pricing_sheet["A" + str(row_start + 3)] = "Material Price"
        pricing_sheet["A" + str(row_start + 3)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 3)] = R_ROOM_PRICING_LIST[i][3]
        pricing_sheet["C" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["D" + str(row_start + 3)] = 0
        pricing_sheet["D" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["D" + str(row_start + 3)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["E" + str(row_start + 3)] = "=D" + str(row_start + 3) + "*C" + str(row_start + 3)
        pricing_sheet["E" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["E" + str(row_start + 3)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 4)] = "Hardware Price"
        pricing_sheet["A" + str(row_start + 4)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 4)] = R_ROOM_PRICING_LIST[i][4]
        pricing_sheet["C" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["D" + str(row_start + 4)] = 0
        pricing_sheet["D" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["D" + str(row_start + 4)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["E" + str(row_start + 4)] = "=D" + str(row_start + 4) + "*C" + str(row_start + 4)
        pricing_sheet["E" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["E" + str(row_start + 4)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 5)] = "Accessories Price"
        pricing_sheet["A" + str(row_start + 5)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 5)] = R_ROOM_PRICING_LIST[i][5]
        pricing_sheet["C" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["D" + str(row_start + 5)] = 0
        pricing_sheet["D" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["D" + str(row_start + 5)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["E" + str(row_start + 5)] = "=D" + str(row_start + 5) + "*C" + str(row_start + 5)
        pricing_sheet["E" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["E" + str(row_start + 5)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 6)] = "Wood Panel Price"
        pricing_sheet["A" + str(row_start + 6)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 6)] = R_ROOM_PRICING_LIST[i][6]
        pricing_sheet["C" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["D" + str(row_start + 6)] = 0
        pricing_sheet["D" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["D" + str(row_start + 6)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["E" + str(row_start + 6)] = "=D" + str(row_start + 6) + "*C" + str(row_start + 6)
        pricing_sheet["E" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["E" + str(row_start + 6)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 7)] = "Room Total Price"
        pricing_sheet["A" + str(row_start + 7)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 7)] = R_ROOM_PRICING_LIST[i][9]
        pricing_sheet["C" + str(row_start + 7)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        pricing_sheet["A" + str(row_start + 9)] = ""
        row_start = pricing_sheet.max_row

    if len(SPECIAL_ORDER_PARTS_LIST) != 0:
        pricing_sheet["B" + str(row_start + 2)] = "Special Order Items (Not Calculated in Project Total)"
        pricing_sheet["B" + str(row_start + 2)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["A" + str(row_start + 3)] = "Special Order Price"
        pricing_sheet["A" + str(row_start + 3)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["B" + str(row_start + 3)] = "See Special Order tab for details"
        pricing_sheet["C" + str(row_start + 3)] = "=SUM('Special Order'!O2:O" + str(len(SPECIAL_ORDER_PARTS_LIST) + 1) + ")"
        pricing_sheet["C" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

    pricing_sheet["B" + str(row_start + 6)] = "Project Totals"
    pricing_sheet["B" + str(row_start + 6)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["A" + str(row_start + 7)] = "Room Count"
    pricing_sheet["A" + str(row_start + 7)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 7)] = str(len(R_ROOM_PRICING_LIST))
    pricing_sheet["A" + str(row_start + 8)] = "Special Order Items Count"
    pricing_sheet["A" + str(row_start + 8)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["B" + str(row_start + 8)] = "See Special Order tab for details"
    pricing_sheet["C" + str(row_start + 8)] = str(len(SPECIAL_ORDER_PARTS_LIST))
    pricing_sheet["A" + str(row_start + 9)] = "Material Price"
    pricing_sheet["A" + str(row_start + 9)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 9)] = sum(map(float, R_PROJECT_TOTAL_MATERIAL))
    pricing_sheet["C" + str(row_start + 9)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 10)] = "Hardware Price"
    pricing_sheet["A" + str(row_start + 10)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 10)] = sum(map(float, R_PROJECT_TOTAL_HARDWARE))
    pricing_sheet["C" + str(row_start + 10)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 11)] = "Accessories Price"
    pricing_sheet["A" + str(row_start + 11)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 11)] = sum(map(float, R_PROJECT_TOTAL_ACCESSORIES))
    pricing_sheet["C" + str(row_start + 11)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 12)] = "Wood Panel Price"
    pricing_sheet["A" + str(row_start + 12)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 12)] = sum(map(float, R_PROJECT_TOTAL_WOOD_PANEL))
    pricing_sheet["C" + str(row_start + 12)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 13)] = "Project Subtotal"
    pricing_sheet["A" + str(row_start + 13)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 13)] = sum(map(float, R_PROJECT_TOTAL_PRICE))
    pricing_sheet["C" + str(row_start + 13)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 14)] = "Delivery Charge (10%)"
    pricing_sheet["A" + str(row_start + 14)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 14)] = "=(" + str((sum(map(float, R_PROJECT_TOTAL_MATERIAL)) + sum(map(float, R_PROJECT_TOTAL_HARDWARE)) + sum(map(float, R_PROJECT_TOTAL_ACCESSORIES)) + sum(map(float, R_PROJECT_TOTAL_WOOD_PANEL)))) + ")*.10"
    pricing_sheet["C" + str(row_start + 14)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 15)] = "Adjusted Project Subtotal"
    pricing_sheet["A" + str(row_start + 15)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 15)] = "=" + str(sum(map(float, R_PROJECT_TOTAL_PRICE))) + "+C" + str(row_start + 14)
    pricing_sheet["C" + str(row_start + 15)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 16)] = "O.D.C Price"
    pricing_sheet["A" + str(row_start + 16)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 16)] = 0
    pricing_sheet["C" + str(row_start + 16)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 17)] = "Tear Out Price"
    pricing_sheet["A" + str(row_start + 17)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 17)] = 0
    pricing_sheet["C" + str(row_start + 17)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 18)] = "Discounts"
    pricing_sheet["A" + str(row_start + 18)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 18)] = "=SUM(E:E)"
    pricing_sheet["C" + str(row_start + 18)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["C" + str(row_start + 18)].font = openpyxl.styles.Font(color="00339966")
    pricing_sheet["A" + str(row_start + 19)] = "Grand Total"
    pricing_sheet["A" + str(row_start + 19)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 19)] = "=" + "C" + str(row_start + 15) + "+C" + str(row_start + 16) + "+C" + str(row_start + 17) + "-C" + str(row_start + 18)
    pricing_sheet["C" + str(row_start + 19)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

    set_column_width(pricing_sheet)
    try:
        wb.save(filename=parts_file)
    except PermissionError:
        print("Retail Pricing Spreadsheet Open...Please close file: " + str(parts_file))
    print("Retail Pricing Summary Created")


def generate_franchise_pricing_summary(parts_file):
    try:
        import openpyxl
        import et_xmlfile
        import pandas
    except ModuleNotFoundError:
        python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
        sys.path.append(python_lib_path)

    wb = openpyxl.Workbook()
    pricing_sheet = wb.active
    pricing_sheet.title = "Franchise Pricing Summary"
    pricing_sheet.HeaderFooter.oddHeader.left.text = "Client Name: {}\nClient ID: {}".format(CLIENT_NAME, CLIENT_ID)
    pricing_sheet.HeaderFooter.oddHeader.center.text = "Pricing Summary Sheet\nJob Number: {}".format(JOB_NUMBER)
    pricing_sheet.HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
    row_start = 0

    for i in range(len(F_ROOM_PRICING_LIST)):
        pricing_sheet["A" + str(row_start + 1)] = "Room Name"
        pricing_sheet["A" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 1)] = F_ROOM_PRICING_LIST[i][0] + " (Franchise)"
        pricing_sheet["C" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["D" + str(row_start + 1)] = R_ROOM_PRICING_LIST[i][0] + " (Retail)"
        pricing_sheet["D" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["E" + str(row_start + 1)] = "Discount %"
        pricing_sheet["E" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["F" + str(row_start + 1)] = "Discount Value"
        pricing_sheet["F" + str(row_start + 1)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["A" + str(row_start + 2)] = "Special Order Items Count"
        pricing_sheet["A" + str(row_start + 2)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["B" + str(row_start + 2)] = "See Special Order tab for details"
        pricing_sheet["C" + str(row_start + 2)] = str(F_ROOM_PRICING_LIST[i][7])
        pricing_sheet["D" + str(row_start + 2)] = str(R_ROOM_PRICING_LIST[i][7])
        pricing_sheet["A" + str(row_start + 3)] = "Material Price"
        pricing_sheet["A" + str(row_start + 3)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 3)] = F_ROOM_PRICING_LIST[i][3]
        pricing_sheet["C" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["D" + str(row_start + 3)] = R_ROOM_PRICING_LIST[i][3]
        pricing_sheet["D" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["E" + str(row_start + 3)] = 0
        pricing_sheet["E" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["E" + str(row_start + 3)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["F" + str(row_start + 3)] = "=E" + str(row_start + 3) + "*C" + str(row_start + 3)
        pricing_sheet["F" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["F" + str(row_start + 3)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 4)] = "Hardware Price"
        pricing_sheet["A" + str(row_start + 4)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 4)] = F_ROOM_PRICING_LIST[i][4]
        pricing_sheet["C" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["D" + str(row_start + 4)] = R_ROOM_PRICING_LIST[i][4]
        pricing_sheet["D" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["E" + str(row_start + 4)] = 0
        pricing_sheet["E" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["E" + str(row_start + 4)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["F" + str(row_start + 4)] = "=E" + str(row_start + 4) + "*C" + str(row_start + 4)
        pricing_sheet["F" + str(row_start + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["F" + str(row_start + 4)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 5)] = "Accessories Price"
        pricing_sheet["A" + str(row_start + 5)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 5)] = F_ROOM_PRICING_LIST[i][5]
        pricing_sheet["C" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["D" + str(row_start + 5)] = R_ROOM_PRICING_LIST[i][5]
        pricing_sheet["D" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["E" + str(row_start + 5)] = 0
        pricing_sheet["E" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["E" + str(row_start + 5)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["F" + str(row_start + 5)] = "=E" + str(row_start + 5) + "*C" + str(row_start + 5)
        pricing_sheet["F" + str(row_start + 5)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["F" + str(row_start + 5)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 6)] = "Wood Panel Price"
        pricing_sheet["A" + str(row_start + 6)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 6)] = F_ROOM_PRICING_LIST[i][6]
        pricing_sheet["C" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["D" + str(row_start + 6)] = R_ROOM_PRICING_LIST[i][6]
        pricing_sheet["D" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["E" + str(row_start + 6)] = 0
        pricing_sheet["E" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_PERCENTAGE
        pricing_sheet["E" + str(row_start + 6)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["F" + str(row_start + 6)] = "=E" + str(row_start + 6) + "*C" + str(row_start + 6)
        pricing_sheet["F" + str(row_start + 6)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["F" + str(row_start + 6)].font = openpyxl.styles.Font(color="00339966")
        pricing_sheet["A" + str(row_start + 7)] = "Room Total Price"
        pricing_sheet["A" + str(row_start + 7)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["C" + str(row_start + 7)] = F_ROOM_PRICING_LIST[i][9]
        pricing_sheet["C" + str(row_start + 7)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["D" + str(row_start + 7)] = R_ROOM_PRICING_LIST[i][9]
        pricing_sheet["D" + str(row_start + 7)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        pricing_sheet["A" + str(row_start + 9)] = ""
        row_start = pricing_sheet.max_row

    if len(SPECIAL_ORDER_PARTS_LIST) != 0:
        pricing_sheet["B" + str(row_start + 2)] = "Special Order Items (Not Calculated in Project Total)"
        pricing_sheet["B" + str(row_start + 2)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["A" + str(row_start + 3)] = "Special Order Price"
        pricing_sheet["A" + str(row_start + 3)].font = openpyxl.styles.Font(bold=True)
        pricing_sheet["B" + str(row_start + 3)] = "See Special Order tab for details"
        pricing_sheet["C" + str(row_start + 3)] = "=SUM('Special Order'!O2:O" + str(len(SPECIAL_ORDER_PARTS_LIST) + 1) + ")"
        pricing_sheet["C" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        pricing_sheet["D" + str(row_start + 3)] = "=SUM('Special Order'!O2:O" + str(len(SPECIAL_ORDER_PARTS_LIST) + 1) + ")"
        pricing_sheet["D" + str(row_start + 3)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

    pricing_sheet["B" + str(row_start + 6)] = "Project Totals"
    pricing_sheet["B" + str(row_start + 6)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 6)] = "Franchise"
    pricing_sheet["C" + str(row_start + 6)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["D" + str(row_start + 6)] = "Retail"
    pricing_sheet["D" + str(row_start + 6)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["A" + str(row_start + 7)] = "Room Count"
    pricing_sheet["A" + str(row_start + 7)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 7)] = str(len(F_ROOM_PRICING_LIST))
    pricing_sheet["D" + str(row_start + 7)] = str(len(R_ROOM_PRICING_LIST))
    pricing_sheet["A" + str(row_start + 8)] = "Special Order Items Count"
    pricing_sheet["A" + str(row_start + 8)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["B" + str(row_start + 8)] = "See Special Order tab for details"
    pricing_sheet["C" + str(row_start + 8)] = str(len(SPECIAL_ORDER_PARTS_LIST))
    pricing_sheet["D" + str(row_start + 8)] = str(len(SPECIAL_ORDER_PARTS_LIST))
    pricing_sheet["A" + str(row_start + 9)] = "Material Price"
    pricing_sheet["A" + str(row_start + 9)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 9)] = sum(map(float, F_PROJECT_TOTAL_MATERIAL))
    pricing_sheet["C" + str(row_start + 9)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["D" + str(row_start + 9)] = sum(map(float, R_PROJECT_TOTAL_MATERIAL))
    pricing_sheet["D" + str(row_start + 9)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 10)] = "Hardware Price"
    pricing_sheet["A" + str(row_start + 10)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 10)] = sum(map(float, F_PROJECT_TOTAL_HARDWARE))
    pricing_sheet["C" + str(row_start + 10)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["D" + str(row_start + 10)] = sum(map(float, R_PROJECT_TOTAL_HARDWARE))
    pricing_sheet["D" + str(row_start + 10)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 11)] = "Accessories Price"
    pricing_sheet["A" + str(row_start + 11)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 11)] = sum(map(float, F_PROJECT_TOTAL_ACCESSORIES))
    pricing_sheet["C" + str(row_start + 11)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["D" + str(row_start + 11)] = sum(map(float, R_PROJECT_TOTAL_ACCESSORIES))
    pricing_sheet["D" + str(row_start + 11)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 12)] = "Wood Panel Price"
    pricing_sheet["A" + str(row_start + 12)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 12)] = sum(map(float, F_PROJECT_TOTAL_WOOD_PANEL))
    pricing_sheet["C" + str(row_start + 12)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["D" + str(row_start + 12)] = sum(map(float, R_PROJECT_TOTAL_WOOD_PANEL))
    pricing_sheet["D" + str(row_start + 12)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 13)] = "Project Subtotal"
    pricing_sheet["A" + str(row_start + 13)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 13)] = sum(map(float, F_PROJECT_TOTAL_PRICE))
    pricing_sheet["C" + str(row_start + 13)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["D" + str(row_start + 13)] = sum(map(float, R_PROJECT_TOTAL_PRICE))
    pricing_sheet["D" + str(row_start + 13)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 14)] = "Delivery Charge (10%)"
    pricing_sheet["A" + str(row_start + 14)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 14)] = 0
    pricing_sheet["C" + str(row_start + 14)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["D" + str(row_start + 14)] = "=(" + str((sum(map(float, R_PROJECT_TOTAL_MATERIAL)) + sum(map(float, R_PROJECT_TOTAL_HARDWARE)) + sum(map(float, R_PROJECT_TOTAL_ACCESSORIES)) + sum(map(float, R_PROJECT_TOTAL_WOOD_PANEL)))) + ")*.10"
    pricing_sheet["D" + str(row_start + 14)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 15)] = "Adjusted Project Subtotal"
    pricing_sheet["A" + str(row_start + 15)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 15)] = "=" + str(sum(map(float, F_PROJECT_TOTAL_PRICE)))
    pricing_sheet["C" + str(row_start + 15)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["D" + str(row_start + 15)] = "=" + str(sum(map(float, R_PROJECT_TOTAL_PRICE))) + "+D" + str(row_start + 14)
    pricing_sheet["D" + str(row_start + 15)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 16)] = "O.D.C Price"
    pricing_sheet["A" + str(row_start + 16)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 16)] = 0
    pricing_sheet["C" + str(row_start + 16)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 17)] = "Tear Out Price"
    pricing_sheet["A" + str(row_start + 17)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 17)] = 0
    pricing_sheet["C" + str(row_start + 17)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["A" + str(row_start + 18)] = "Discounts"
    pricing_sheet["A" + str(row_start + 18)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 18)] = "=SUM(E:E)"
    pricing_sheet["C" + str(row_start + 18)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["C" + str(row_start + 18)].font = openpyxl.styles.Font(color="00339966")
    pricing_sheet["A" + str(row_start + 19)] = "Grand Total"
    pricing_sheet["A" + str(row_start + 19)].font = openpyxl.styles.Font(bold=True)
    pricing_sheet["C" + str(row_start + 19)] = "=" + "C" + str(row_start + 15) + "+C" + str(row_start + 16) + "+C" + str(row_start + 17) + "-C" + str(row_start + 18)
    pricing_sheet["C" + str(row_start + 19)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
    pricing_sheet["D" + str(row_start + 19)] = "=" + "D" + str(row_start + 15) + "+D" + str(row_start + 16) + "+D" + str(row_start + 17) + "-D" + str(row_start + 18)
    pricing_sheet["D" + str(row_start + 19)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

    set_column_width(pricing_sheet)
    try:
        wb.save(filename=parts_file)
    except PermissionError:
        print("Pricing Spreadsheet Open...Please close file: " + str(parts_file))
    print("Franchise Pricing Summary Created")


def generate_parts_summary(parts_file, materials_sheet, hardware_sheet, accessories_sheet, wood_panel_sheet, so_sheet):
    try:
        import pandas
        import numpy 
    except ModuleNotFoundError:
        python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
        sys.path.append(python_lib_path)

    with pandas.ExcelWriter(parts_file, mode='a') as writer:
        if materials_sheet is not None:
            df_materials = pandas.read_excel(parts_file, sheet_name='Materials').query('"BB" <= SKU_NUMBER < "EB" or "GL" <= SKU_NUMBER <= "PM~"')
            # df_materials = pandas.read_excel(parts_file, sheet_name='Materials').loc[(df['SKU_NUMBER']>="BB") & (df['SKU_NUMBER']<"EB")]
            # df.loc[(df['Salary_in_1000']>=100) & (df['Age']< 60) & (df['FT_Team'].str.startswith('S')),['Name','FT_Team']]
            materials_summary = pandas.pivot_table(df_materials, index=['ROOM_NAME', 'WALL_NAME', 'MATERIAL', 'PART_NAME', 'PART_DIMENSIONS', 'THICKNESS'], values=['QUANTITY'], aggfunc=numpy.sum)
            materials_summary.to_excel(writer, sheet_name='Materials Summary')
            writer.sheets['Materials Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nClient ID: {}".format(CLIENT_NAME, CLIENT_ID)
            writer.sheets['Materials Summary'].HeaderFooter.oddHeader.center.text = "Materials Summary Sheet\nJob Number: {}".format(JOB_NUMBER)
            writer.sheets['Materials Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
            set_column_width(writer.sheets['Materials Summary'])

        if hardware_sheet is not None:
            df_hardware = pandas.read_excel(parts_file, sheet_name='Hardware')
            hardware_summary = pandas.pivot_table(df_hardware, index=['ROOM_NAME', 'VENDOR_NAME', 'VENDOR_ITEM', 'PART_NAME'], values='QUANTITY', aggfunc=numpy.sum)
            hardware_summary.to_excel(writer, sheet_name='Hardware Summary')
            writer.sheets['Hardware Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nClient ID: {}".format(CLIENT_NAME, CLIENT_ID)
            writer.sheets['Hardware Summary'].HeaderFooter.oddHeader.center.text = "Hardware Summary Sheet\nJob Number: {}".format(JOB_NUMBER)
            writer.sheets['Hardware Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
            set_column_width(writer.sheets['Hardware Summary'])
            
        if accessories_sheet is not None:
            df_accessories = pandas.read_excel(parts_file, sheet_name='Accessories')
            accessories_summary = pandas.pivot_table(df_accessories, index=['ROOM_NAME', 'VENDOR_NAME', 'VENDOR_ITEM', 'PART_NAME', 'LENGTH'], values='QUANTITY', aggfunc=numpy.sum)
            accessories_summary.to_excel(writer, sheet_name='Accessories Summary')
            writer.sheets['Accessories Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nClient ID: {}".format(CLIENT_NAME, CLIENT_ID)
            writer.sheets['Accessories Summary'].HeaderFooter.oddHeader.center.text = "Accessories Summary Sheet\nJob Number: {}".format(JOB_NUMBER)
            writer.sheets['Accessories Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
            set_column_width(writer.sheets['Accessories Summary'])

        if wood_panel_sheet is not None:
            df_wood_panel = pandas.read_excel(parts_file, sheet_name='Wood Panel')
            wood_panel_summary = pandas.pivot_table(df_wood_panel, index=['ROOM_NAME', 'STAIN_COLOR', 'PART_NAME', 'PART_DIMENSIONS', 'THICKNESS'], values=['QUANTITY'], aggfunc=numpy.sum)
            wood_panel_summary.to_excel(writer, sheet_name='Wood Panel Summary')
            writer.sheets['Wood Panel Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nClient ID: {}".format(CLIENT_NAME, CLIENT_ID)
            writer.sheets['Wood Panel Summary'].HeaderFooter.oddHeader.center.text = "Wood Panel Summary Sheet\nJob Number: {}".format(JOB_NUMBER)
            writer.sheets['Wood Panel Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
            set_column_width(writer.sheets['Wood Panel Summary'])

        if so_sheet is not None:
            df_so = pandas.read_excel(parts_file, sheet_name='Special Order')
            so_summary = pandas.pivot_table(df_so, index=['ROOM_NAME', 'MATERIAL', 'PART_NAME', 'PART_DIMENSIONS', 'THICKNESS'], values=['QUANTITY'], aggfunc=numpy.sum)
            so_summary.to_excel(writer, sheet_name='Special Order Summary')
            writer.sheets['Special Order Summary'].HeaderFooter.oddHeader.left.text = "Client Name: {}\nClient ID: {}".format(CLIENT_NAME, CLIENT_ID)
            writer.sheets['Special Order Summary'].HeaderFooter.oddHeader.center.text = "Special Order Summary Sheet\nJob Number: {}".format(JOB_NUMBER)
            writer.sheets['Special Order Summary'].HeaderFooter.oddHeader.right.text = "Project Name: {}\nDesign Date: {}".format(PROJECT_NAME, DESIGN_DATE)
            set_column_width(writer.sheets['Special Order Summary'])

    print("Parts Summary Created")
    display_parts_summary(parts_file)


def generate_retail_parts_list():
    props = bpy.context.window_manager.sn_project
    proj = props.get_project()
    cleaned_name = proj.get_clean_name(proj.name)
    project_dir = bpy.context.preferences.addons['snap'].preferences.project_dir
    selected_project = os.path.join(project_dir, cleaned_name)
    parts_file = os.path.join(selected_project, "Retail_Pricing_Parts_List" + "(" + str(cleaned_name) + ").xlsx")

    print("Creating Retail Pricing Summary...")
    generate_retail_pricing_summary(parts_file)

    if not os.path.exists(project_dir):
        print("Projects Directory does not exist")
    
    # Start by opening the spreadsheet and selecting the main sheet
    try:
        import openpyxl
        import et_xmlfile
    except ModuleNotFoundError:
        python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
        sys.path.append(python_lib_path)

    wb = openpyxl.load_workbook(parts_file)
    sheet1 = None
    sheet2 = None
    sheet3 = None
    sheet4 = None
    sheet5 = None

    if len(MATERIAL_PARTS_LIST) != 0:
        sheet1 = wb.create_sheet()
        sheet1.title = "Materials"
        sheet1.append(["ROOM_NAME", "WALL_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "RETAIL_PRICE", "LABOR", "CALCULATED_PRICE"])

        for i in range(len(MATERIAL_PARTS_LIST)):
            sheet1["A" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][0]                                          #ROOM_NAME
            sheet1["B" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][17]                                         #WALL_NAME
            sheet1["C" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][1]                                          #SKU_NUMBER 
            sheet1["D" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][2]                                          #VENDOR_NAME
            sheet1["E" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][3]                                          #VENDOR_ITEM
            sheet1["F" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][4]                                          #PART LABELID
            sheet1["G" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][5]                                          #MATERIAL
            sheet1["H" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][6]                                          #PART_NAME    
            sheet1["I" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][7]                                          #QUANTITY
            sheet1["J" + str((i + 1) + 1)] = str(MATERIAL_PARTS_LIST[i][8]) + " x " + str(MATERIAL_PARTS_LIST[i][9])      #PART_DIMENSIONS           
            sheet1["K" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][10]                                         #THICKNESS
            sheet1["N" + str((i + 1) + 1)] = float(MATERIAL_PARTS_LIST[i][11])                                  #RETAIL_PRICE
            sheet1["O" + str((i + 1) + 1)] = float(MATERIAL_PARTS_LIST[i][12])                                  #LABOR
            
            if MATERIAL_PARTS_LIST[i][1][:2] in material_types:
                if 'SF' in MATERIAL_PARTS_LIST[i][13]:
                    sheet1["L" + str((i + 1) + 1)] = get_square_footage(float(MATERIAL_PARTS_LIST[i][8]), float(MATERIAL_PARTS_LIST[i][9]))
                    sheet1["P" + str((i + 1) + 1)] = (((float(MATERIAL_PARTS_LIST[i][11]) * int(MATERIAL_PARTS_LIST[i][7])) * get_square_footage(float(MATERIAL_PARTS_LIST[i][8]), float(MATERIAL_PARTS_LIST[i][9])))  + float(MATERIAL_PARTS_LIST[i][12]))   #CALCULATED_PRICE
                    if len(EDGEBANDING) > 0 and i < len(EDGEBANDING):
                        if MATERIAL_PARTS_LIST[i][4] == EDGEBANDING[i][0]:
                            sheet1["Q" + str((i + 1) + 1)] = str(EDGEBANDING[i][1]) + " " + str(EDGEBANDING[i][2])
                if 'LF' in MATERIAL_PARTS_LIST[i][13]:
                    if MATERIAL_PARTS_LIST[i][14] is not None:
                        eb_length = get_eb_measurements(MATERIAL_PARTS_LIST[i][14], float(MATERIAL_PARTS_LIST[i][8]), float(MATERIAL_PARTS_LIST[i][9]))
                    sheet1["M" + str((i + 1) + 1)] = str(get_linear_footage(eb_length)) + " (" + str(MATERIAL_PARTS_LIST[i][14]) + ")"
                    sheet1["P" + str((i + 1) + 1)] = (float(MATERIAL_PARTS_LIST[i][11]) * int(MATERIAL_PARTS_LIST[i][7])) * get_linear_footage(eb_length)   #CALCULATED_PRICE
            
            sheet1["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet1["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet1["P" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet1.max_row
        sheet1["P" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet1["P" + str(row_max + 3)] = "TOTAL"
        sheet1["P" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet1["P" + str(row_max + 4)] = "=SUM(P2:P" + str(row_max) + ")"

        set_column_width(sheet1)
    else:
        print("Material Parts List Empty")

    if len(HARDWARE_PARTS_LIST) != 0:
        sheet2 = wb.create_sheet()
        sheet2.title = "Hardware"
        sheet2.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "", "QUANTITY", "RETAIL_PRICE", "CALCULATED_PRICE"])
        
        for i in range(len(HARDWARE_PARTS_LIST)):
            sheet2["A" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][0]         
            sheet2["B" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][1]
            sheet2["C" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][2]           
            sheet2["D" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][3]
            sheet2["E" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][4] 
            sheet2["F" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][5]
            sheet2["H" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][6]
            sheet2["I" + str((i + 1) + 1)] = float(HARDWARE_PARTS_LIST[i][7])
            sheet2["I" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet2["J" + str((i + 1) + 1)] = float(HARDWARE_PARTS_LIST[i][7]) * int(HARDWARE_PARTS_LIST[i][6])
            sheet2["J" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet2.max_row
        sheet2["J" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet2["J" + str(row_max + 3)] = "TOTAL"
        sheet2["J" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet2["J" + str(row_max + 4)] = "=SUM(J2:J" + str(row_max) + ")"

        set_column_width(sheet2)
    else:
        print("Hardware Parts List Empty")

    if len(ACCESSORY_PARTS_LIST) != 0:
        sheet3 = wb.create_sheet()
        sheet3.title = "Accessories"
        sheet3.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "LENGTH", "QUANTITY", "RETAIL_PRICE", "CALCULATED_PRICE"])

        for i in range(len(ACCESSORY_PARTS_LIST)):
            sheet3["A" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][0]     
            sheet3["B" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][1]
            sheet3["C" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][2]           
            sheet3["D" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][3]
            sheet3["E" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][4]           
            sheet3["F" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][5]
            if ACCESSORY_PARTS_LIST[i][6] == 0:
                sheet3["G" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][6]
            else:
                sheet3["G" + str((i + 1) + 1)] = str(ACCESSORY_PARTS_LIST[i][6]) + " (" + str(get_linear_footage(float(ACCESSORY_PARTS_LIST[i][6]))) + "LF)"
            sheet3["H" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][7]
            sheet3["I" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet3["I" + str((i + 1) + 1)] = float(ACCESSORY_PARTS_LIST[i][8])
            sheet3["J" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

            if float(ACCESSORY_PARTS_LIST[i][6]) > 0.0:
                sheet3["J" + str((i + 1) + 1)] = (float(ACCESSORY_PARTS_LIST[i][8]) * int(ACCESSORY_PARTS_LIST[i][7])) * get_linear_footage(float(ACCESSORY_PARTS_LIST[i][6]))
            else:
                sheet3["J" + str((i + 1) + 1)] = float(ACCESSORY_PARTS_LIST[i][8]) * int(ACCESSORY_PARTS_LIST[i][7])

        row_max = sheet3.max_row
        sheet3["J" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet3["J" + str(row_max + 3)] = "TOTAL"
        sheet3["J" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet3["J" + str(row_max + 4)] = "=SUM(J2:J" + str(row_max) + ")"

        set_column_width(sheet3)
    else:
        print("Accessory Parts List Empty")

    if len(WOOD_PANEL_PARTS_LIST) != 0:
        sheet4 = wb.create_sheet()
        sheet4.title = "Wood Panel"
        sheet4.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "STAIN_COLOR", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "RETAIL_PRICE", "LABOR", "CALCULATED_PRICE"])

        for i in range(len(WOOD_PANEL_PARTS_LIST)):
            sheet4["A" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][0]                                          #ROOM_NAME
            sheet4["B" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][1]                                          #SKU_NUMBER 
            sheet4["C" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][2]                                          #VENDOR_NAME
            sheet4["D" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][3]                                          #VENDOR_ITEM
            sheet4["E" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][4]                                          #PART LABELID
            sheet4["F" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][5]                                          #STAIN_COLOR
            sheet4["G" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][6]                                          #PART_NAME    
            sheet4["H" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][7]                                          #QUANTITY
            sheet4["I" + str((i + 1) + 1)] = str(WOOD_PANEL_PARTS_LIST[i][8]) + " x " + str(WOOD_PANEL_PARTS_LIST[i][9])           #PART_DIMENSIONS           
            sheet4["J" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][10]                                         #THICKNESS
            sheet4["K" + str((i + 1) + 1)] = get_square_footage(float(WOOD_PANEL_PARTS_LIST[i][8]), float(WOOD_PANEL_PARTS_LIST[i][9]))
            sheet4["M" + str((i + 1) + 1)] = float(WOOD_PANEL_PARTS_LIST[i][11])                                  #RETAIL_PRICE
            sheet4["M" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet4["N" + str((i + 1) + 1)] = float(WOOD_PANEL_PARTS_LIST[i][12])                                  #LABOR
            sheet4["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet4["O" + str((i + 1) + 1)] = float(WOOD_PANEL_PARTS_LIST[i][11])  #CALCULATED_PRICE
            sheet4["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet4.max_row
        sheet4["O" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet4["O" + str(row_max + 3)] = "TOTAL"
        sheet4["O" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet4["O" + str(row_max + 4)] = "=SUM(O2:O" + str(row_max) + ")"

        set_column_width(sheet4)
    else:
        print("Wood Panel Parts List Empty")

    if len(SPECIAL_ORDER_PARTS_LIST) != 0:
        sheet5 = wb.create_sheet()
        sheet5.title = "Special Order"
        sheet5.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "RETAIL_PRICE", "LABOR", "CALCULATED_PRICE"])

        for i in range(len(SPECIAL_ORDER_PARTS_LIST)):
            sheet5["A" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][0]                                          #ROOM_NAME
            sheet5["B" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][1]                                          #SKU_NUMBER 
            sheet5["C" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][2]                                          #VENDOR_NAME
            sheet5["D" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][3]                                          #VENDOR_ITEM
            sheet5["E" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][4]                                          #PART LABELID
            sheet5["F" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][5]                                          #MATERIAL
            sheet5["G" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][6]                                          #PART_NAME    
            sheet5["H" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][7]                                          #QUANTITY
            sheet5["I" + str((i + 1) + 1)] = str(SPECIAL_ORDER_PARTS_LIST[i][8]) + " x " + str(SPECIAL_ORDER_PARTS_LIST[i][9])      #PART_DIMENSIONS           
            sheet5["J" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][10]                                         #THICKNESS
            sheet5["M" + str((i + 1) + 1)] = float(SPECIAL_ORDER_PARTS_LIST[i][11])                                  #RETAIL_PRICE
            sheet5["M" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet5["N" + str((i + 1) + 1)] = float(SPECIAL_ORDER_PARTS_LIST[i][12])                                  #LABOR
            sheet5["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet5["O" + str((i + 1) + 1)] = (float(SPECIAL_ORDER_PARTS_LIST[i][11]) + float(SPECIAL_ORDER_PARTS_LIST[i][12])) * int(SPECIAL_ORDER_PARTS_LIST[i][7])
            sheet5["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet5.max_row
        sheet5["O" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet5["O" + str(row_max + 3)] = "TOTAL"
        sheet5["O" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet5["O" + str(row_max + 4)] = "=SUM(O2:O" + str(row_max) + ")"

        set_column_width(sheet5)
    else:
        print("Special Order Parts List Empty")

    # Save the spreadsheet
    wb.save(filename=parts_file)
    print("Retail Pricing Parts List Generated")
    print("Creating Retail Parts Summary...")
    generate_parts_summary(parts_file, sheet1, sheet2, sheet3, sheet4, sheet5)


def generate_franchise_parts_list():
    props = bpy.context.window_manager.sn_project
    proj = props.get_project()
    cleaned_name = proj.get_clean_name(proj.name)
    project_dir = bpy.context.preferences.addons['snap'].preferences.project_dir
    selected_project = os.path.join(project_dir, cleaned_name)
    parts_file = os.path.join(selected_project, "Franchise_Pricing_Parts_List" + "(" + str(cleaned_name) + ").xlsx")

    print("Creating Franchise Pricing Summary...")
    generate_franchise_pricing_summary(parts_file)

    if not os.path.exists(project_dir):
        print("Projects Directory does not exist")
    
    # Start by opening the spreadsheet and selecting the main sheet
    try:
        import openpyxl
        import et_xmlfile
    except ModuleNotFoundError:
        python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
        sys.path.append(python_lib_path)

    wb = openpyxl.load_workbook(parts_file)
    sheet1 = None
    sheet2 = None
    sheet3 = None
    sheet4 = None
    sheet5 = None

    if len(MATERIAL_PARTS_LIST) != 0:
        sheet1 = wb.create_sheet()
        sheet1.title = "Materials"
        sheet1.append(["ROOM_NAME", "WALL_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "FRANCHISE_PRICE", "LABOR", "CALCULATED_PRICE"])

        for i in range(len(MATERIAL_PARTS_LIST)):
            sheet1["A" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][0]                                          #ROOM_NAME
            sheet1["B" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][17]                                         #WALL_NAME
            sheet1["C" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][1]                                          #SKU_NUMBER 
            sheet1["D" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][2]                                          #VENDOR_NAME
            sheet1["E" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][3]                                          #VENDOR_ITEM
            sheet1["F" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][4]                                          #PART LABELID
            sheet1["G" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][5]                                          #MATERIAL
            sheet1["H" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][6]                                          #PART_NAME    
            sheet1["I" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][7]                                          #QUANTITY
            sheet1["J" + str((i + 1) + 1)] = str(MATERIAL_PARTS_LIST[i][8]) + " x " + str(MATERIAL_PARTS_LIST[i][9])      #PART_DIMENSIONS           
            sheet1["K" + str((i + 1) + 1)] = MATERIAL_PARTS_LIST[i][10]                                         #THICKNESS
            sheet1["N" + str((i + 1) + 1)] = float(MATERIAL_PARTS_LIST[i][16])                                  #RETAIL_PRICE
            sheet1["O" + str((i + 1) + 1)] = float(MATERIAL_PARTS_LIST[i][15])                                  #LABOR
            
            if MATERIAL_PARTS_LIST[i][1][:2] in material_types:
                if 'SF' in MATERIAL_PARTS_LIST[i][13]:
                    sheet1["L" + str((i + 1) + 1)] = get_square_footage(float(MATERIAL_PARTS_LIST[i][8]), float(MATERIAL_PARTS_LIST[i][9]))
                    sheet1["P" + str((i + 1) + 1)] = (((float(MATERIAL_PARTS_LIST[i][16]) * int(MATERIAL_PARTS_LIST[i][7])) * get_square_footage(float(MATERIAL_PARTS_LIST[i][8]), float(MATERIAL_PARTS_LIST[i][9])))  + float(MATERIAL_PARTS_LIST[i][15]))   #CALCULATED_PRICE
                if 'LF' in MATERIAL_PARTS_LIST[i][13]:
                    if MATERIAL_PARTS_LIST[i][14] is not None:
                        eb_length = get_eb_measurements(MATERIAL_PARTS_LIST[i][14], float(MATERIAL_PARTS_LIST[i][8]), float(MATERIAL_PARTS_LIST[i][9]))
                    sheet1["M" + str((i + 1) + 1)] = str(get_linear_footage(eb_length)) + " (" + str(MATERIAL_PARTS_LIST[i][14]) + ")"
                    sheet1["P" + str((i + 1) + 1)] = (float(MATERIAL_PARTS_LIST[i][16]) * int(MATERIAL_PARTS_LIST[i][7])) * get_linear_footage(eb_length)   #CALCULATED_PRICE
            
            sheet1["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet1["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet1["P" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet1.max_row
        sheet1["P" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet1["P" + str(row_max + 3)] = "TOTAL"
        sheet1["P" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet1["P" + str(row_max + 4)] = "=SUM(P2:P" + str(row_max) + ")"

        set_column_width(sheet1)
    else:
        print("Material Parts List Empty")

    if len(HARDWARE_PARTS_LIST) != 0:
        sheet2 = wb.create_sheet()
        sheet2.title = "Hardware"
        sheet2.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "", "QUANTITY", "FRANCHISE_PRICE", "CALCULATED_PRICE"])
        
        for i in range(len(HARDWARE_PARTS_LIST)):
            sheet2["A" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][0]         
            sheet2["B" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][1]
            sheet2["C" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][2]           
            sheet2["D" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][3]
            sheet2["E" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][4] 
            sheet2["F" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][5]
            sheet2["H" + str((i + 1) + 1)] = HARDWARE_PARTS_LIST[i][6]
            sheet2["I" + str((i + 1) + 1)] = float(HARDWARE_PARTS_LIST[i][8])
            sheet2["I" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet2["J" + str((i + 1) + 1)] = float(HARDWARE_PARTS_LIST[i][8]) * int(HARDWARE_PARTS_LIST[i][6])
            sheet2["J" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet2.max_row
        sheet2["J" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet2["J" + str(row_max + 3)] = "TOTAL"
        sheet2["J" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet2["J" + str(row_max + 4)] = "=SUM(J2:J" + str(row_max) + ")"

        set_column_width(sheet2)
    else:
        print("Hardware Parts List Empty")

    if len(ACCESSORY_PARTS_LIST) != 0:
        sheet3 = wb.create_sheet()
        sheet3.title = "Accessories"
        sheet3.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "PART_NAME", "LENGTH", "QUANTITY", "FRANCHISE_PRICE", "CALCULATED_PRICE"])

        for i in range(len(ACCESSORY_PARTS_LIST)):
            sheet3["A" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][0]     
            sheet3["B" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][1]
            sheet3["C" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][2]           
            sheet3["D" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][3]
            sheet3["E" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][4]           
            sheet3["F" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][5]
            if ACCESSORY_PARTS_LIST[i][6] == 0:
                sheet3["G" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][6]
            else:
                sheet3["G" + str((i + 1) + 1)] = str(ACCESSORY_PARTS_LIST[i][6]) + " (" + str(get_linear_footage(float(ACCESSORY_PARTS_LIST[i][6]))) + "LF)"
            sheet3["H" + str((i + 1) + 1)] = ACCESSORY_PARTS_LIST[i][7]
            sheet3["I" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet3["I" + str((i + 1) + 1)] = float(ACCESSORY_PARTS_LIST[i][9])
            sheet3["J" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

            if float(ACCESSORY_PARTS_LIST[i][6]) > 0.0:
                sheet3["J" + str((i + 1) + 1)] = (float(ACCESSORY_PARTS_LIST[i][9]) * int(ACCESSORY_PARTS_LIST[i][7])) * get_linear_footage(float(ACCESSORY_PARTS_LIST[i][6]))
            else:
                sheet3["J" + str((i + 1) + 1)] = float(ACCESSORY_PARTS_LIST[i][9]) * int(ACCESSORY_PARTS_LIST[i][7])

        row_max = sheet3.max_row
        sheet3["J" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet3["J" + str(row_max + 3)] = "TOTAL"
        sheet3["J" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet3["J" + str(row_max + 4)] = "=SUM(J2:J" + str(row_max) + ")"

        set_column_width(sheet3)
    else:
        print("Accessory Parts List Empty")

    if len(WOOD_PANEL_PARTS_LIST) != 0:
        sheet4 = wb.create_sheet()
        sheet4.title = "Wood Panel"
        sheet4.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "STAIN_COLOR", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "FRANCHISE_PRICE", "LABOR", "CALCULATED_PRICE"])

        for i in range(len(WOOD_PANEL_PARTS_LIST)):
            sheet4["A" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][0]                                          #ROOM_NAME
            sheet4["B" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][1]                                          #SKU_NUMBER 
            sheet4["C" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][2]                                          #VENDOR_NAME
            sheet4["D" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][3]                                          #VENDOR_ITEM
            sheet4["E" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][4]                                          #PART LABELID
            sheet4["F" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][5]                                          #STAIN_COLOR
            sheet4["G" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][6]                                          #PART_NAME    
            sheet4["H" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][7]                                          #QUANTITY
            sheet4["I" + str((i + 1) + 1)] = str(WOOD_PANEL_PARTS_LIST[i][8]) + " x " + str(WOOD_PANEL_PARTS_LIST[i][9])           #PART_DIMENSIONS           
            sheet4["J" + str((i + 1) + 1)] = WOOD_PANEL_PARTS_LIST[i][10]                                         #THICKNESS
            sheet4["K" + str((i + 1) + 1)] = get_square_footage(float(WOOD_PANEL_PARTS_LIST[i][8]), float(WOOD_PANEL_PARTS_LIST[i][9]))
            sheet4["M" + str((i + 1) + 1)] = float(WOOD_PANEL_PARTS_LIST[i][14])                                  #FRANCHISE_PRICE
            sheet4["M" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet4["N" + str((i + 1) + 1)] = float(WOOD_PANEL_PARTS_LIST[i][13])                                  #LABOR
            sheet4["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet4["O" + str((i + 1) + 1)] = float(WOOD_PANEL_PARTS_LIST[i][14])  #CALCULATED_PRICE
            sheet4["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet4.max_row
        sheet4["O" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet4["O" + str(row_max + 3)] = "TOTAL"
        sheet4["O" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet4["O" + str(row_max + 4)] = "=SUM(O2:O" + str(row_max) + ")"

        set_column_width(sheet4)
    else:
        print("Wood Panel Parts List Empty")

    if len(SPECIAL_ORDER_PARTS_LIST) != 0:
        sheet5 = wb.create_sheet()
        sheet5.title = "Special Order"
        sheet5.append(["ROOM_NAME", "SKU_NUMBER", "VENDOR_NAME", "VENDOR_ITEM", "PART LABELID", "MATERIAL", "PART_NAME", "QUANTITY",
                        "PART_DIMENSIONS", "THICKNESS", "SQUARE_FT", "LINEAR_FT", "FRANCHISE_PRICE", "LABOR", "CALCULATED_PRICE"])

        for i in range(len(SPECIAL_ORDER_PARTS_LIST)):
            sheet5["A" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][0]                                          #ROOM_NAME
            sheet5["B" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][1]                                          #SKU_NUMBER 
            sheet5["C" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][2]                                          #VENDOR_NAME
            sheet5["D" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][3]                                          #VENDOR_ITEM
            sheet5["E" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][4]                                          #PART LABELID
            sheet5["F" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][5]                                          #MATERIAL
            sheet5["G" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][6]                                          #PART_NAME    
            sheet5["H" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][7]                                          #QUANTITY
            sheet5["I" + str((i + 1) + 1)] = str(SPECIAL_ORDER_PARTS_LIST[i][8]) + " x " + str(SPECIAL_ORDER_PARTS_LIST[i][9])      #PART_DIMENSIONS           
            sheet5["J" + str((i + 1) + 1)] = SPECIAL_ORDER_PARTS_LIST[i][10]                                         #THICKNESS
            sheet5["M" + str((i + 1) + 1)] = float(SPECIAL_ORDER_PARTS_LIST[i][14])                                  #FRANCHISE_PRICE
            sheet5["M" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet5["N" + str((i + 1) + 1)] = float(SPECIAL_ORDER_PARTS_LIST[i][13])                                  #LABOR
            sheet5["N" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
            sheet5["O" + str((i + 1) + 1)] = (float(SPECIAL_ORDER_PARTS_LIST[i][14]) + float(SPECIAL_ORDER_PARTS_LIST[i][13])) * int(SPECIAL_ORDER_PARTS_LIST[i][7])
            sheet5["O" + str((i + 1) + 1)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE

        row_max = sheet5.max_row
        sheet5["O" + str(row_max + 3)].font = openpyxl.styles.Font(bold=True)
        sheet5["O" + str(row_max + 3)] = "TOTAL"
        sheet5["O" + str(row_max + 4)].number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE
        sheet5["O" + str(row_max + 4)] = "=SUM(O2:O" + str(row_max) + ")"

        set_column_width(sheet5)
    else:
        print("Special Order Parts List Empty")

    # Save the spreadsheet
    wb.save(filename=parts_file)
    print("Franchise Pricing Parts List Generated")
    print("Creating Franchise Parts Summary...")
    generate_parts_summary(parts_file, sheet1, sheet2, sheet3, sheet4, sheet5)


def get_labor_costs(part_name):
    if part_name is not None:
        part_name = part_name.upper()
    rows = sn_db.query_db(
        "SELECT\
            SKU,\
            DisplayName,\
            FranchisePrice,\
            RetailPrice\
        FROM\
            {CCItems}\
        WHERE\
            ProductType == 'LBR' AND\
            DisplayName LIKE 'LABOR - {}'\
        ;".format(part_name, CCItems="CCItems_" + bpy.context.preferences.addons['snap'].preferences.franchise_location)
    )
    if len(rows) == 0:
        retail_price = 0
        franchise_price = 0
        # print("No Labor Price Found for {}".format(part_name))
    for row in rows:
        sku = row[0]
        display_name = row[1]
        franchise_price = float(row[2])
        retail_price = float(row[3])
        # print("SKU: {}  DisplayName: {}  Part Name: {}  RetailPrice: {}".format(sku, display_name, part_name, retail_price))

    return retail_price, franchise_price


def get_eb_measurements(eb_orientation, length, width):
    if length > width:
        long_side = length
        short_side = width
    else:
        long_side = width
        short_side = length

    if eb_orientation == 'S1' or eb_orientation == 'S2':
        return short_side
    elif eb_orientation == 'L1' or eb_orientation == 'L2':
        return long_side
    else:
        return long_side


def calculate_wood_panel_price(style_name, is_glaze, glaze_style, glaze_color):
    print("Wood Panel")


def get_price_by_sku(sku_num):
    sku_num = str(sku_num)
    rows = sn_db.query_db(
        "SELECT\
            SKU,\
            DisplayName,\
            RetailPrice,\
            FranchisePrice,\
            UOM,\
            VendorName,\
            VendorItemNum\
        FROM\
            {CCItems}\
        WHERE\
            SKU = '{SKU}'\
        ;".format(CCItems="CCItems_" + bpy.context.preferences.addons['snap'].preferences.franchise_location, SKU=sku_num)
    )
    if len(rows) == 0:
        retail_price = 0
        franchise_price = 0
        uom = ''
        display_name = ''
        vendor_name = ''
        vendor_item = 0
        print("No Prices Returned for SKU:  " + sku_num)
    for row in rows:
        display_name = row[1]
        retail_price = float(row[2])
        franchise_price = float(row[3])
        uom = row[4]
        vendor_name = row[5]
        vendor_item = row[6]

    return retail_price, franchise_price

def get_pricing_info(sku_num, qty, length_inches=0.0, width_inches=0.0, style_name=None, is_glaze=False, glaze_style=None, glaze_color=None, part_name=None, eb_orientation=None):
    length_inches = float(length_inches)
    width_inches = float(width_inches)
    wood_panel_pricing_file = os.path.join(sn_paths.ROOT_DIR, "db_init", "Wood_Panel_Pricing.xlsx")
    style_type = ''
    glaze_price = 0.0
    r_labor_price = 0
    f_labor_price = 0
    eb_length = 0
    rows = sn_db.query_db(
        "SELECT\
            SKU,\
            Name,\
            RetailPrice,\
            FranchisePrice,\
            UOM,\
            VendorName,\
            VendorItemNum\
        FROM\
            {CCItems}\
        WHERE\
            SKU = '{SKU}'\
        ;".format(CCItems="CCItems_" + bpy.context.preferences.addons['snap'].preferences.franchise_location, SKU=sku_num)
    )
    if len(rows) == 0:
        retail_price = 0
        franchise_price = 0
        uom = ''
        name = ''
        vendor_name = ''
        vendor_item = 0
        print("No Prices Returned for SKU:  " + sku_num)
    for row in rows:
        name = row[1]
        retail_price = float(row[2])
        franchise_price = float(row[3])
        uom = row[4]
        vendor_name = row[5]
        vendor_item = row[6]

    price_check(sku_num, franchise_price, retail_price)

    if part_name is not None:
        # Get Labor Costs
        if sku_num[:2] in material_types and sku_num[:2] not in 'EB':
            labor = get_labor_costs(part_name)
            if 'Drawer Bottom' in part_name:
                if 'BB' in sku_num[:2]:
                    labor = get_labor_costs('BB DRAWER INSET BOTTOM')
                else:
                    labor = get_labor_costs('MEL DRAWER CAP BOTTOM')
            R_LABOR_PRICES.append(labor[0] * int(qty))
            F_LABOR_PRICES.append(labor[1] * int(qty))
            r_labor_price = labor[0]
            f_labor_price = labor[1]
    if 'SF' in uom or 'SQFT' in uom and not sku_num[:2] in hardware_types and not sku_num[:2] in accessory_types:
        R_MATERIAL_PRICES.append((((get_square_footage(length_inches, width_inches)) * retail_price) + r_labor_price) * int(qty))
        R_MATERIAL_SQUARE_FOOTAGE.append(get_square_footage(length_inches, width_inches))
        F_MATERIAL_PRICES.append((((get_square_footage(length_inches, width_inches)) * franchise_price) + f_labor_price) * int(qty))
        F_MATERIAL_SQUARE_FOOTAGE.append(get_square_footage(length_inches, width_inches))
    if 'LF' in uom and not sku_num[:2] in hardware_types and not sku_num[:2] in accessory_types:
        if eb_orientation is not None:
            eb_length = get_eb_measurements(eb_orientation, length_inches, width_inches)
        R_MATERIAL_PRICES.append((get_linear_footage(eb_length) * retail_price) * int(qty))
        R_MATERIAL_LINEAR_FOOTAGE.append(get_linear_footage(eb_length))
        F_MATERIAL_PRICES.append((get_linear_footage(eb_length) * franchise_price) * int(qty))
        F_MATERIAL_LINEAR_FOOTAGE.append(get_linear_footage(eb_length))
    if 'PL' or 'SN' in sku_num[:2]:
        r_stain_total = 0
        r_door_total = 0
        f_stain_total = 0
        f_door_total = 0
        veneer_price = get_price_by_sku('VN-0000002')
        three_quarter_mdf_price = get_price_by_sku('WD-0000010')
        quarter_mdf_price = get_price_by_sku('WD-0000007')
        wood_labor_price = get_price_by_sku('LB-0000014')
        primer_price = 0
        sealer_price = get_price_by_sku('CH-0000001')
        topcoat_price = get_price_by_sku('CH-0000002')
        catalyst_price = get_price_by_sku('CH-0000003')
        jacketboard_price = get_price_by_sku('WD-0000002')
        care_price = get_price_by_sku('CH-0000004')
        glaze_labor = 0
    
        if style_name is not None:
            try:
                import pandas
            except ModuleNotFoundError:
                python_lib_path = os.path.join(sn_paths.ROOT_DIR, "python_lib")
                sys.path.append(python_lib_path)

            wood_panel_data_frame = pandas.read_excel(wood_panel_pricing_file, header=None, skiprows=1, usecols=[0, 1, 2, 3, 4, 5], names=['STYLE', 'TYPE', 'PANEL', 'SKU', 'AP_SKU', 'GL_NUM'])

            if 'door'.upper() in style_name.upper():
                style_type = 'Door'
            else:
                style_type = 'Drawer'

            for index, data_row in wood_panel_data_frame.iterrows():
                if data_row['STYLE'] in style_name and style_type == data_row['TYPE']:
                    panel = data_row['PANEL']
                    molding_sku = data_row['SKU']
                    ap_molding_sku = data_row['AP_SKU']
                    glazing_lines = data_row['GL_NUM']
                    lf_molding = round(((length_inches / 12) + (width_inches / 12)) * 2, 2)
                    lf_glaze = round(((length_inches / 12) + (width_inches / 12)) * 2, 2) * glazing_lines
                    sf_center_panel = get_square_footage(length_inches-4, width_inches-4)
                    sf_door = get_square_footage(length_inches, width_inches)
                    molding_price = get_price_by_sku(molding_sku)
                    ap_molding_price = get_price_by_sku(ap_molding_sku)
                    
                    if 'SN' in sku_num[:2]:
                        if is_glaze:
                            glaze_labor = get_labor_costs(glaze_style)
                            r_stain_total = (retail_price * sf_door) + (sealer_price[0] * sf_door) + (lf_glaze * glaze_labor[0]) + ((topcoat_price[0] + catalyst_price[0]) * sf_door)
                            f_stain_total = (franchise_price * sf_door) + (sealer_price[1] * sf_door) + (lf_glaze * glaze_labor[1]) + ((topcoat_price[1] + catalyst_price[1]) * sf_door)
                        else:
                            r_stain_total = (retail_price * sf_door) + (sealer_price[0] * sf_door) + ((topcoat_price[0] + catalyst_price[0]) * sf_door)
                            f_stain_total = (franchise_price * sf_door) + (sealer_price[1] * sf_door) + ((topcoat_price[1] + catalyst_price[1]) * sf_door)
                        if 'Raised' in panel:
                            r_door_total = (lf_molding * molding_price[0]) + (lf_molding * ap_molding_price[0]) + (sf_center_panel * jacketboard_price[0]) + wood_labor_price[0]
                            f_door_total = (lf_molding * molding_price[1]) + (lf_molding * ap_molding_price[1]) + (sf_center_panel * jacketboard_price[1]) + wood_labor_price[1]
                        if 'Recessed' in panel:
                            r_door_total = (lf_molding * molding_price[0]) + (lf_molding * ap_molding_price[0]) + (sf_center_panel * veneer_price[0]) + wood_labor_price[0]
                            f_door_total = (lf_molding * molding_price[1]) + (lf_molding * ap_molding_price[1]) + (sf_center_panel * veneer_price[1]) + wood_labor_price[1]

                    if 'PL' in sku_num[:2]:
                        if 'PL-0000004' or 'PL-0000005' or 'PL-0000009' in sku_num:
                            primer_price = get_price_by_sku('CH-0000005')
                        else:
                            primer_price = get_price_by_sku('CH-0000006')
                        if is_glaze:
                            glaze_labor = get_labor_costs(glaze_style)
                            r_stain_total = (primer_price[0] * sf_door) + (retail_price * sf_door) + (sealer_price[0] * sf_door) + (lf_glaze * glaze_labor[0]) + ((topcoat_price[0] + catalyst_price[0]) * sf_door)
                            f_stain_total = (primer_price[1] * sf_door) + (franchise_price * sf_door) + (sealer_price[1] * sf_door) + (lf_glaze * glaze_labor[1]) + ((topcoat_price[1] + catalyst_price[1]) * sf_door)
                        else:   
                            r_stain_total = (primer_price[0] * sf_door) + ((retail_price + care_price[0]) * sf_door)
                            f_stain_total = (primer_price[1] * sf_door) + ((franchise_price + care_price[1]) * sf_door)
                        if 'Raised' in panel:
                            r_door_total = (lf_molding * molding_price[0]) + (lf_molding * ap_molding_price[0]) + (sf_center_panel * three_quarter_mdf_price[0]) + wood_labor_price[0]
                            f_door_total = (lf_molding * molding_price[1]) + (lf_molding * ap_molding_price[1]) + (sf_center_panel * three_quarter_mdf_price[1]) + wood_labor_price[1]
                        else:
                            r_door_total = (lf_molding * molding_price[0]) + (lf_molding * ap_molding_price[0]) + (sf_center_panel * quarter_mdf_price[0]) + wood_labor_price[0]
                            f_door_total = (lf_molding * molding_price[1]) + (lf_molding * ap_molding_price[1]) + (sf_center_panel * quarter_mdf_price[1]) + wood_labor_price[1]


                    R_LABOR_PRICES.append(wood_labor_price[0] * int(qty))
                    F_LABOR_PRICES.append(wood_labor_price[1] * int(qty))
                    R_WOOD_PANEL_PRICES.append((r_door_total + r_stain_total) * int(qty))
                    F_WOOD_PANEL_PRICES.append((f_door_total + f_stain_total) * int(qty))
                    retail_price = r_door_total + r_stain_total
                    franchise_price = f_door_total + f_stain_total
            # if 'Melamine' in style_name:
            #     retail_price = retail_price + glaze_price
            #     franchise_price = franchise_price + glaze_price
            #     r_rpd_price = get_square_footage(length_inches, width_inches) * retail_price
            #     f_rpd_price = get_square_footage(length_inches, width_inches) * franchise_price
            #     labor = get_labor_costs('5PC MEL DOOR & DRWR FACE')
            #     r_labor_price = labor[0]
            #     f_labor_price = labor[1]
            #     R_LABOR_PRICES.append(r_labor_price * int(qty))
            #     R_WOOD_PANEL_PRICES.append((r_rpd_price + r_labor_price) * int(qty))
            #     F_LABOR_PRICES.append(f_labor_price * int(qty))
            #     F_WOOD_PANEL_PRICES.append((f_rpd_price + f_labor_price) * int(qty))
    if sku_num[:2] in hardware_types:
        R_HARDWARE_PRICES.append(retail_price * int(qty))
        F_HARDWARE_PRICES.append(franchise_price * int(qty))
    if sku_num[:2] in accessory_types:
        if 'LF' in uom:
            R_ACCESSORY_PRICES.append(get_linear_footage(length_inches) * (retail_price * int(qty)))
            F_ACCESSORY_PRICES.append(get_linear_footage(length_inches) * (franchise_price * int(qty)))
        else:
            R_ACCESSORY_PRICES.append(retail_price * int(qty))
            F_ACCESSORY_PRICES.append(franchise_price * int(qty))
    if sku_num[:2] in special_order_types:
        R_SPECIAL_ORDER_PRICES.append((retail_price * int(qty)) + r_labor_price)
        F_SPECIAL_ORDER_PRICES.append((franchise_price * int(qty)) + f_labor_price)

    return str(retail_price), str(franchise_price), name, vendor_name, vendor_item, style_name, r_labor_price, f_labor_price, uom


def calculate_project_price(xml_file):
    tree = None
    root = None
    eb_orientation = None
    eb_counter = 1

    if os.path.exists(xml_file):
        tree = ET.parse(xml_file)
        root = tree.getroot().find("Job")
    else:
        attrib = dict([ (k, v) for k, v in ns.items() ])
        root = ET.Element('Batch', attrib)
        tree = ET.ElementTree(root)

    name = root.find("Name")
    var = root.findall("Var")
    set_job_info(name, var)

    mfg = root.find("Manufacturing")
                                
    #Collect and Output
    for item in root.findall("Item"):

        # for item_name in item.findall("Name"):

        for description in item.findall("Description"):
            DESCRIPTION = description.text

        for part in item.findall("Part"):
            PART_LABEL_ID = part.attrib.get('LabelID')

            for quantity in part.findall("Quantity"):
                QUANTITY = quantity.text

            for part_name in part.findall("Name"):
                if part_name.text == "Top Cleat" or part_name.text == "Bottom Cleat":
                    PART_NAME = "Cleat"
                elif part_name.text == "Left Drawer Side" or part_name.text == "Right Drawer Side":
                    PART_NAME = "Drawer Side"
                elif part_name.text == "Left Door" or part_name.text == "Right Door":
                    PART_NAME = "Door"
                else:
                    PART_NAME = part_name.text

                for width in part.findall("Width"):
                    WIDTH = width.text

                for length in part.findall("Length"):
                    LENGTH = length.text

                for thickness in part.findall("Thickness"):
                    THICKNESS = thickness.text

            for label in mfg.findall("Label"):
                label_id = label.attrib.get("ID", None)
                # part_id = part.attrib.get("PartID", None)
                eb1 = None
                eb2 = None
                eb3 = None
                eb4 = None
                rod_length = None
                style_name = 'None'
                glaze_style = 'None'
                glaze_color = 'None'
                is_glaze = False
                glass_color = 'None'
                wall_name = None
                if PART_LABEL_ID == label_id:
                    # Use the iterator to deal with repeated children and still in order
                    label_iterator = iter(label)
                    for _ in range(int(len(label) / 3)):
                        sku_value = None
                        NAME = next(label_iterator).text
                        TYPE = next(label_iterator).text
                        VALUE = next(label_iterator).text
                        if 'sku' in NAME:
                            sku_value = VALUE
                        if NAME == 'wallname':
                            wall_name = VALUE
                        if NAME == 'edgeband1':
                            eb1 = VALUE
                        if NAME == 'edgeband2':
                            eb2 = VALUE
                        if NAME == 'edgeband3':
                            eb3 = VALUE
                        if NAME == 'edgeband4':
                            eb4 = VALUE
                        if 'rod'.upper() in PART_NAME.upper():
                            if NAME == 'lenx':
                                rod_length = VALUE
                        if NAME == 'style':
                            style_name = VALUE
                        if NAME == 'glazecolor':
                            glaze_color = VALUE
                            if glaze_color != 'None':
                                is_glaze = True
                        if NAME == 'glazestyle':
                            glaze_style = VALUE
                            if glaze_style != 'None':
                                is_glaze = True
                        if NAME == 'glasscolor':
                            glass_color = VALUE
                    # for sku_value in label.findall("Value"):
                        if sku_value is not None:
                            if sku_value[:2] in material_types:
                                SKU_NUMBER = sku_value
                                print(SKU_NUMBER)
                                if sku_value[:2] in 'EB':
                                    if eb1 is not None and eb_counter == 1:
                                        eb_orientation = eb1
                                        eb_counter = 2
                                        print(eb_orientation)
                                        S_COUNT.append(eb_orientation)
                                    elif eb2 is not None and eb_counter == 2 or eb_counter == 1:
                                        eb_orientation = eb2
                                        eb_counter = 3
                                        print(eb_orientation)
                                        L_COUNT.append(eb_orientation)
                                    elif eb3 is not None and eb_counter == 3:
                                        eb_orientation = eb3
                                        eb_counter = 4
                                        print(eb_orientation)
                                        S_COUNT.append(eb_orientation)
                                    elif eb4 is not None and eb_counter == 4:
                                        eb_orientation = eb4
                                        eb_counter = 1
                                        print(eb_orientation)
                                        L_COUNT.append(eb_orientation)
                                # Continue statement to skip over edgebanding for glass shelves
                                if 'Glass' in PART_NAME and sku_value[:2] in 'EB':
                                    continue
                                # Continue statement to skip over edgebanding for Wood Panel doors
                                if style_name != 'None' and sku_value[:2] in 'EB':
                                    continue
                                if style_name != 'None' and sku_value[:2] not in 'EB':
                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME)
                                    if is_glaze:
                                        if glaze_color != 'None' and glaze_style != 'None':
                                            PART_NAME = PART_NAME + " (" + style_name + ", Glazed " + glaze_color + "(" + glaze_style + ")"
                                        else:
                                            PART_NAME = PART_NAME + " (" + style_name + ", Glazed " + glaze_color + ")"
                                    else:
                                        PART_NAME = PART_NAME + " (" + style_name + ")"
                                    WOOD_PANEL_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1]])
                                else:
                                    if 'Glass' in PART_NAME and glass_color is not None and "None" not in glass_color:
                                        PART_NAME = PART_NAME + " (" + glass_color + ")"
                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME, eb_orientation)
                                    MATERIAL_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[8], eb_orientation, pricing_info[7], pricing_info[1], wall_name])
                            if sku_value[:2] in hardware_types:
                                SKU_NUMBER = sku_value
                                pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY)
                                PART_NAME = pricing_info[2]
                                HARDWARE_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, QUANTITY, pricing_info[0], pricing_info[1]])
                            if sku_value[:2] in accessory_types:
                                SKU_NUMBER = sku_value
                                if 'rod'.upper() in PART_NAME.upper():
                                    LENGTH = rod_length
                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH)
                                else:
                                    LENGTH = 0
                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY)
                                if 'pull'.upper() in pricing_info[2] or 'kn'.upper() in pricing_info[2]:
                                    PART_NAME = pricing_info[2][:-10]
                                else:
                                    PART_NAME = pricing_info[2]
                                ACCESSORY_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, LENGTH, QUANTITY, pricing_info[0], pricing_info[1]])
                            if sku_value[:2] in special_order_types:
                                if 'Glass' in PART_NAME and glass_color is not None and "None" not in glass_color:
                                    PART_NAME = PART_NAME + " (" + glass_color + ")"
                                SKU_NUMBER = sku_value
                                pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME)
                                SPECIAL_ORDER_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1]])
            eb_counter = 1
            EDGEBANDING.append((label_id, len(S_COUNT), len(L_COUNT)))
            S_COUNT.clear()
            L_COUNT.clear()

        for assembly in item.findall("Assembly"):

            for part in assembly.findall("Part"):
                DESCRIPTION = description.text
                PART_LABEL_ID = part.attrib.get('LabelID')

                for quantity in part.findall("Quantity"):
                    QUANTITY = quantity.text

                for part_name in part.findall("Name"):
                    if part_name.text == "Top Cleat" or part_name.text == "Bottom Cleat":
                        PART_NAME = "Cleat"
                    elif part_name.text == "Left Drawer Side" or part_name.text == "Right Drawer Side":
                        PART_NAME = "Drawer Side"
                    elif part_name.text == "Left Door" or part_name.text == "Right Door":
                        PART_NAME = "Door"
                    else:
                        PART_NAME = part_name.text

                    if part_name.text is not None:
                        for width in part.findall("Width"):
                            WIDTH = width.text

                        for length in part.findall("Length"):
                            LENGTH = length.text

                        for thickness in part.findall("Thickness"):
                            THICKNESS = thickness.text

                for label in mfg.findall("Label"):
                    label_id = label.attrib.get("ID", None)
                    # part_id = part.attrib.get("PartID", None)
                    eb1 = None
                    eb2 = None
                    eb3 = None
                    eb4 = None
                    style_name = 'None'
                    glaze_style = 'None'
                    glaze_color = 'None'
                    is_glaze = False
                    glass_color = 'None'
                    wall_name = None
                    if PART_LABEL_ID == label_id:
                        # Use the iterator to deal with repeated children and still in order
                        label_iterator = iter(label)
                        for _ in range(int(len(label) / 3)):
                            sku_value = None
                            NAME = next(label_iterator).text
                            TYPE = next(label_iterator).text
                            VALUE = next(label_iterator).text
                            if 'sku' in NAME:
                                sku_value = VALUE
                            if NAME == 'wallname':
                                wall_name = VALUE
                            if NAME == 'edgeband1':
                                eb1 = VALUE
                            if NAME == 'edgeband2':
                                eb2 = VALUE
                            if NAME == 'edgeband3':
                                eb3 = VALUE
                            if NAME == 'edgeband4':
                                eb4 = VALUE
                            if 'rod'.upper() in PART_NAME.upper():
                                if NAME == 'lenx':
                                    rod_length = VALUE
                            if NAME == 'style':
                                style_name = VALUE
                                if 'Aviano' in style_name:
                                    print(style_name)
                            if NAME == 'glazecolor':
                                glaze_color = VALUE
                                if glaze_color != 'None':
                                    is_glaze = True
                            if NAME == 'glazestyle':
                                glaze_style = VALUE
                                if glaze_style != 'None':
                                    is_glaze = True
                            if NAME == 'glasscolor':
                                glass_color = VALUE
                        # for sku_value in label.findall("Value"):
                            if sku_value is not None:
                                if sku_value[:2] in material_types:
                                    SKU_NUMBER = sku_value
                                    print(SKU_NUMBER)
                                    if sku_value[:2] in 'EB':
                                        if eb1 is not None and eb_counter == 1:
                                            eb_orientation = eb1
                                            eb_counter = 2
                                            print(eb_orientation)
                                            S_COUNT.append(eb_orientation)
                                        elif eb2 is not None and eb_counter == 2 or eb_counter == 1:
                                            eb_orientation = eb2
                                            eb_counter = 3
                                            print(eb_orientation)
                                            L_COUNT.append(eb_orientation)
                                        elif eb3 is not None and eb_counter == 3:
                                            eb_orientation = eb3
                                            eb_counter = 4
                                            print(eb_orientation)
                                            S_COUNT.append(eb_orientation)
                                        elif eb4 is not None and eb_counter == 4:
                                            eb_orientation = eb4
                                            eb_counter = 1
                                            print(eb_orientation)
                                            L_COUNT.append(eb_orientation)
                                    # Continue statement to skip over edgebanding for glass shelves
                                    if 'Glass' in PART_NAME and sku_value[:2] in 'EB':
                                        continue
                                    # Continue statement to skip over edgebanding for Wood Panel doors
                                    if style_name != 'None' and sku_value[:2] in 'EB':
                                        continue
                                    if style_name != 'None' and sku_value[:2] not in 'EB':
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME)
                                        if is_glaze:
                                            if glaze_color != 'None' and glaze_style != 'None':
                                                PART_NAME = PART_NAME + " (" + style_name + ", Glazed " + glaze_color + "(" + glaze_style + ")"
                                            else:
                                                PART_NAME = PART_NAME + " (" + style_name + ", Glazed " + glaze_color + ")"
                                        else:
                                            PART_NAME = PART_NAME + " (" + style_name + ")"
                                        WOOD_PANEL_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1]])
                                    else:
                                        if 'Glass' in PART_NAME and glass_color is not None and "None" not in glass_color:
                                            PART_NAME = PART_NAME + " (" + glass_color + ")"
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME, eb_orientation)
                                        MATERIAL_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[8], eb_orientation, pricing_info[7], pricing_info[1], wall_name])
                                if sku_value[:2] in hardware_types:
                                    SKU_NUMBER = sku_value
                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY)
                                    PART_NAME = pricing_info[2]
                                    HARDWARE_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, QUANTITY, pricing_info[0], pricing_info[1]])
                                if sku_value[:2] in accessory_types:
                                    SKU_NUMBER = sku_value
                                    if 'rod'.upper() in PART_NAME.upper():
                                        LENGTH = rod_length
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH)
                                    else:
                                        LENGTH = 0
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY)
                                    if 'pull'.upper() in pricing_info[2] or 'kn'.upper() in pricing_info[2]:
                                        PART_NAME = pricing_info[2][:-10]
                                    else:
                                        PART_NAME = pricing_info[2]
                                    ACCESSORY_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, LENGTH, QUANTITY, pricing_info[0], pricing_info[1]])
                                if sku_value[:2] in special_order_types:
                                    if 'Glass' in PART_NAME and glass_color is not None and "None" not in glass_color:
                                        PART_NAME = PART_NAME + " (" + glass_color + ")"
                                    SKU_NUMBER = sku_value
                                    pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME)
                                    SPECIAL_ORDER_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1]])
                eb_counter = 1
                EDGEBANDING.append((label_id, len(S_COUNT), len(L_COUNT)))
                S_COUNT.clear()
                L_COUNT.clear()

            for assembly in assembly.findall("Assembly"):

                for part in assembly.findall("Part"):
                    DESCRIPTION = description.text
                    PART_LABEL_ID = part.attrib.get('LabelID')

                    for quantity in part.findall("Quantity"):
                        QUANTITY = quantity.text

                    for part_name in part.findall("Name"):
                        if part_name.text == "Top Cleat" or part_name.text == "Bottom Cleat":
                            PART_NAME = "Cleat"
                        elif part_name.text == "Left Drawer Side" or part_name.text == "Right Drawer Side":
                            PART_NAME = "Drawer Side"
                        elif part_name.text == "Left Door" or part_name.text == "Right Door":
                            PART_NAME = "Door"
                        else:
                            PART_NAME = part_name.text

                        if part_name.text is not None:
                            for width in part.findall("Width"):
                                WIDTH = width.text

                            for length in part.findall("Length"):
                                LENGTH = length.text

                            for thickness in part.findall("Thickness"):
                                THICKNESS = thickness.text

                    for label in mfg.findall("Label"):
                        label_id = label.attrib.get("ID", None)
                        # part_id = part.attrib.get("PartID", None)
                        eb1 = None
                        eb2 = None
                        eb3 = None
                        eb4 = None
                        rod_length = None
                        style_name = 'None'
                        glaze_style = 'None'
                        glaze_color = 'None'
                        is_glaze = False
                        glass_color = 'None'
                        wall_name = None
                        if PART_LABEL_ID == label_id:
                            # Use the iterator to deal with repeated children and still in order
                            label_iterator = iter(label)
                            for _ in range(int(len(label) / 3)):
                                sku_value = None
                                NAME = next(label_iterator).text
                                TYPE = next(label_iterator).text
                                VALUE = next(label_iterator).text
                                if 'sku' in NAME:
                                    sku_value = VALUE
                                if NAME == 'wallname':
                                    wall_name = VALUE
                                if NAME == 'edgeband1':
                                    eb1 = VALUE
                                if NAME == 'edgeband2':
                                    eb2 = VALUE
                                if NAME == 'edgeband3':
                                    eb3 = VALUE
                                if NAME == 'edgeband4':
                                    eb4 = VALUE
                                if 'rod'.upper() in PART_NAME.upper():
                                    if NAME == 'lenx':
                                        rod_length = VALUE
                                if NAME == 'style':
                                    style_name = VALUE
                                    if 'Aviano' in style_name:
                                        print(style_name)
                                if NAME == 'glazecolor':
                                    glaze_color = VALUE
                                    if glaze_color != 'None':
                                        is_glaze = True
                                if NAME == 'glazestyle':
                                    glaze_style = VALUE
                                    if glaze_style != 'None':
                                        is_glaze = True
                                if NAME == 'glasscolor':
                                    glass_color = VALUE
                            # for sku_value in label.findall("Value"):
                                if sku_value is not None:
                                    if sku_value[:2] in material_types:
                                        SKU_NUMBER = sku_value
                                        if sku_value[:2] in 'EB':
                                            if eb1 is not None and eb_counter == 1:
                                                eb_orientation = eb1
                                                eb_counter = 2
                                                print(eb_orientation)
                                                S_COUNT.append(eb_orientation)
                                            elif eb2 is not None and eb_counter == 2 or eb_counter == 1:
                                                eb_orientation = eb2
                                                eb_counter = 3
                                                print(eb_orientation)
                                                L_COUNT.append(eb_orientation)
                                            elif eb3 is not None and eb_counter == 3:
                                                eb_orientation = eb3
                                                eb_counter = 4
                                                print(eb_orientation)
                                                S_COUNT.append(eb_orientation)
                                            elif eb4 is not None and eb_counter == 4:
                                                eb_orientation = eb4
                                                eb_counter = 1
                                                print(eb_orientation)
                                                L_COUNT.append(eb_orientation)
                                        # Continue statement to skip over edgebanding for glass shelves
                                        if 'Glass' in PART_NAME and sku_value[:2] in 'EB':
                                            continue
                                        # Continue statement to skip over edgebanding for Wood Panel doors
                                        if style_name != 'None' and sku_value[:2] in 'EB':
                                            continue
                                        if style_name != 'None' and sku_value[:2] not in 'EB':
                                            pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, style_name, is_glaze, glaze_style, glaze_color, PART_NAME)
                                            if is_glaze:
                                                if glaze_color != 'None' and glaze_style != 'None':
                                                    PART_NAME = PART_NAME + " (" + style_name + ", Glazed " + glaze_color + "(" + glaze_style + ")"
                                                else:
                                                    PART_NAME = PART_NAME + " (" + style_name + ", Glazed " + glaze_color + ")"
                                            else:
                                                PART_NAME = PART_NAME + " (" + style_name + ")"
                                            WOOD_PANEL_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1]])
                                        else:
                                            if 'Glass' in PART_NAME and glass_color is not None and "None" not in glass_color:
                                                PART_NAME = PART_NAME + " (" + glass_color + ")"
                                            pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME, eb_orientation)
                                            MATERIAL_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[8], eb_orientation, pricing_info[7], pricing_info[1], wall_name])
                                    if sku_value[:2] in hardware_types:
                                        SKU_NUMBER = sku_value
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY)
                                        PART_NAME = pricing_info[2]
                                        HARDWARE_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, QUANTITY, pricing_info[0], pricing_info[1]])
                                    if sku_value[:2] in accessory_types:
                                        SKU_NUMBER = sku_value
                                        if 'rod'.upper() in PART_NAME.upper():
                                            LENGTH = rod_length
                                            pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH)
                                        else:
                                            LENGTH = 0
                                            pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY)
                                        if 'pull'.upper() in pricing_info[2] or 'kn'.upper() in pricing_info[2]:
                                            PART_NAME = pricing_info[2][:-10]
                                        else:
                                            PART_NAME = pricing_info[2]
                                        ACCESSORY_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, PART_NAME, LENGTH, QUANTITY, pricing_info[0], pricing_info[1]])
                                    if sku_value[:2] in special_order_types:
                                        if 'Glass' in PART_NAME and glass_color is not None and "None" not in glass_color:
                                            PART_NAME = PART_NAME + " (" + glass_color + ")"
                                        SKU_NUMBER = sku_value
                                        pricing_info = get_pricing_info(SKU_NUMBER, QUANTITY, LENGTH, WIDTH, None, False, None, None, PART_NAME)
                                        SPECIAL_ORDER_PARTS_LIST.append([DESCRIPTION, SKU_NUMBER, pricing_info[3], pricing_info[4], PART_LABEL_ID, pricing_info[2], PART_NAME, QUANTITY, LENGTH, WIDTH, THICKNESS, pricing_info[0], pricing_info[6], pricing_info[7], pricing_info[1]])
                    eb_counter = 1
                    EDGEBANDING.append((label_id, len(S_COUNT), len(L_COUNT)))
                    S_COUNT.clear()
                    L_COUNT.clear()

        R_ROOM_TOTAL_PRICE = sum(map(float, R_MATERIAL_PRICES)) + sum(map(float, R_HARDWARE_PRICES)) + sum(map(float, R_ACCESSORY_PRICES)) + sum(map(float, R_WOOD_PANEL_PRICES))
        R_ROOM_PRICING_LIST.append([DESCRIPTION, sum(map(float, R_MATERIAL_SQUARE_FOOTAGE)), sum(map(float, R_MATERIAL_LINEAR_FOOTAGE)), sum(map(float, R_MATERIAL_PRICES)), sum(map(float, R_HARDWARE_PRICES)), sum(map(float, R_ACCESSORY_PRICES)), sum(map(float, R_WOOD_PANEL_PRICES)), len(R_SPECIAL_ORDER_PRICES), sum(map(float, R_LABOR_PRICES)), R_ROOM_TOTAL_PRICE])
        R_PROJECT_TOTAL_HARDWARE.append(sum(map(float, R_HARDWARE_PRICES)))
        R_PROJECT_TOTAL_ACCESSORIES.append(sum(map(float, R_ACCESSORY_PRICES)))
        R_PROJECT_TOTAL_SQUARE_FOOTAGE.append(sum(map(float, R_MATERIAL_SQUARE_FOOTAGE)))
        R_PROJECT_TOTAL_LINEAR_FOOTAGE.append(sum(map(float, R_MATERIAL_LINEAR_FOOTAGE)))
        R_PROJECT_TOTAL_MATERIAL.append(sum(map(float, R_MATERIAL_PRICES)))
        R_PROJECT_TOTAL_WOOD_PANEL.append(sum(map(float, R_WOOD_PANEL_PRICES)))
        R_PROJECT_TOTAL_LABOR.append(sum(map(float, R_LABOR_PRICES)))
        R_PROJECT_TOTAL_PRICE.append(R_ROOM_TOTAL_PRICE)
        R_ROOM_TOTAL_PRICE = 0
        R_MATERIAL_PRICES.clear()
        R_HARDWARE_PRICES.clear()
        R_ACCESSORY_PRICES.clear()
        R_WOOD_PANEL_PRICES.clear()
        R_SPECIAL_ORDER_PRICES.clear()
        R_LABOR_PRICES.clear()
        R_MATERIAL_SQUARE_FOOTAGE.clear()
        R_MATERIAL_LINEAR_FOOTAGE.clear()

        F_ROOM_TOTAL_PRICE = sum(map(float, F_MATERIAL_PRICES)) + sum(map(float, F_HARDWARE_PRICES)) + sum(map(float, F_ACCESSORY_PRICES)) + sum(map(float, F_WOOD_PANEL_PRICES))

        F_ROOM_PRICING_LIST.append([DESCRIPTION, sum(map(float, F_MATERIAL_SQUARE_FOOTAGE)), sum(map(float, F_MATERIAL_LINEAR_FOOTAGE)), sum(map(float, F_MATERIAL_PRICES)), sum(map(float, F_HARDWARE_PRICES)), sum(map(float, F_ACCESSORY_PRICES)), sum(map(float, F_WOOD_PANEL_PRICES)), len(F_SPECIAL_ORDER_PRICES), sum(map(float, F_LABOR_PRICES)), F_ROOM_TOTAL_PRICE])
        F_PROJECT_TOTAL_HARDWARE.append(sum(map(float, F_HARDWARE_PRICES)))
        F_PROJECT_TOTAL_ACCESSORIES.append(sum(map(float, F_ACCESSORY_PRICES)))
        F_PROJECT_TOTAL_SQUARE_FOOTAGE.append(sum(map(float, F_MATERIAL_SQUARE_FOOTAGE)))
        F_PROJECT_TOTAL_LINEAR_FOOTAGE.append(sum(map(float, F_MATERIAL_LINEAR_FOOTAGE)))
        F_PROJECT_TOTAL_MATERIAL.append(sum(map(float, F_MATERIAL_PRICES)))
        F_PROJECT_TOTAL_WOOD_PANEL.append(sum(map(float, F_WOOD_PANEL_PRICES)))
        F_PROJECT_TOTAL_LABOR.append(sum(map(float, F_LABOR_PRICES)))
        F_PROJECT_TOTAL_PRICE.append(F_ROOM_TOTAL_PRICE)
        F_ROOM_TOTAL_PRICE = 0
        F_MATERIAL_PRICES.clear()
        F_HARDWARE_PRICES.clear()
        F_ACCESSORY_PRICES.clear()
        F_WOOD_PANEL_PRICES.clear()
        F_SPECIAL_ORDER_PRICES.clear()
        F_LABOR_PRICES.clear()
        F_MATERIAL_SQUARE_FOOTAGE.clear()
        F_MATERIAL_LINEAR_FOOTAGE.clear()

    print("Calculate Price COMPLETE")
    bpy.context.window.cursor_set("DEFAULT")


class SNAP_OT_Calculate_Price(Operator):
    bl_idname = PRICING_PROPERTY_NAMESPACE + ".calculate_price"
    bl_label = "Calculate Price"
    bl_description = "Calculate Price for Project"

    tmp_filename = "export_temp.py"
    xml_filename = "snap_job.xml"
    proj_dir: StringProperty(name="Project Directory", subtype='DIR_PATH')

    def invoke(self, context, event):
        wm = context.window_manager
        return wm.invoke_props_dialog(self, width=300)

    def draw(self, context):
        layout = self.layout
        props = context.window_manager.sn_project
        proj = props.get_project()
        layout.label(text="Project: {}".format(proj.name))
        box = layout.box()

        for room in proj.rooms:
            col = box.column(align=True)
            row = col.row()
            row.prop(room, "selected", text="")
            row.label(text=room.name)

        row = layout.row()
        row.operator(
            "sn_project_pricing.select_all_rooms", text="Select All", icon='CHECKBOX_HLT').select_all = True
        row.operator(
            "sn_project_pricing.select_all_rooms", text="Deselect All", icon='CHECKBOX_DEHLT').select_all = False

    def create_prep_script(self):
        nrm_dir = self.proj_dir.replace("\\", "/")
        file = open(os.path.join(bpy.app.tempdir, self.tmp_filename), 'w')
        file.write("import bpy\n")
        file.write("bpy.ops.sn_export.export_xml('INVOKE_DEFAULT', xml_path='{}')\n".format(nrm_dir))
        file.close()
        return os.path.join(bpy.app.tempdir, self.tmp_filename)
        
    def execute(self, context):
        bpy.ops.wm.save_mainfile()
        bpy.context.window.cursor_set("WAIT")
        debug_mode = context.preferences.addons["snap"].preferences.debug_mode
        debug_mac = context.preferences.addons["snap"].preferences.debug_mac
        proj_props = bpy.context.window_manager.sn_project
        proj_name = proj_props.projects[proj_props.project_index].name
        path = os.path.join(sn_xml.get_project_dir(), proj_name, self.xml_filename)
        proj = proj_props.projects[proj_props.project_index]

        if os.path.exists(path):
            os.remove(path)

        self.proj_dir = os.path.join(sn_xml.get_project_dir(), proj_name)
        script_path = self.create_prep_script()

        # Call blender in background and run XML export on each room file in project
        for room in proj.rooms:
            if room.selected:
                subprocess.call(bpy.app.binary_path + ' "' + room.file_path + '" -b --python "' + script_path + '"')

        if debug_mode and debug_mac:
            bpy.ops.wm.open_mainfile(filepath=bpy.data.filepath)
            if "Machining" in bpy.data.collections:
                for obj in bpy.data.collections["Machining"].objects:
                    obj.display_type = 'WIRE'

        props = get_pricing_props()
        props.reset()
        props.calculate_price(context)
        return {'FINISHED'}


class SNAP_OT_Select_All_Rooms_Pricing(Operator):
    bl_idname = "sn_project_pricing.select_all_rooms"
    bl_label = "Select All Rooms"
    bl_description = "This will select all of the rooms in the project"

    select_all: BoolProperty(name="Select All", default=True)

    @classmethod
    def poll(cls, context):
        return True

    def execute(self, context):
        props = context.window_manager.sn_project
        proj = props.projects[props.project_index]

        for room in proj.rooms:
            room.selected = self.select_all

        return{'FINISHED'}


class SNAP_PROPS_Pricing(bpy.types.PropertyGroup):

    pricing_tabs: EnumProperty(
        name="Pricing Tabs",
        items=[('RETAIL', "Retail Pricing", 'View Pricing Information'), 
               ('FRANCHISE', "Franchise Pricing", 'View Pricing information')])
    
    # auto_calculate_on_save: BoolProperty(name="Auto Calculate On Save",
    #                                       default=False,
    #                                       description="Automatically calculate the pricing information")

    export_pricing_parts_list: BoolProperty(name="Export Pricing Parts List (.xlsx)",
                                          default=False,
                                          description="Export a list of all parts being priced")
    
    def reset(self):
        '''
            This function resets all of the pricing properties back to their default
            This should be run before you calculate the price to ensure there is no
            values stored from a previous calculation
        '''
        R_MATERIAL_PRICES.clear()
        R_HARDWARE_PRICES.clear()
        R_ACCESSORY_PRICES.clear()
        R_WOOD_PANEL_PRICES.clear()
        R_SPECIAL_ORDER_PRICES.clear()
        R_LABOR_PRICES.clear()
        R_MATERIAL_SQUARE_FOOTAGE.clear()
        R_MATERIAL_LINEAR_FOOTAGE.clear()
        R_PROJECT_TOTAL_HARDWARE.clear()
        R_PROJECT_TOTAL_ACCESSORIES.clear()
        R_PROJECT_TOTAL_MATERIAL.clear()
        R_PROJECT_TOTAL_WOOD_PANEL.clear()
        R_PROJECT_TOTAL_SQUARE_FOOTAGE.clear()
        R_PROJECT_TOTAL_LINEAR_FOOTAGE.clear()
        R_PROJECT_TOTAL_PRICE.clear()
        R_ROOM_PRICING_LIST.clear()

        MATERIAL_PARTS_LIST.clear()
        ACCESSORY_PARTS_LIST.clear()
        HARDWARE_PARTS_LIST.clear()
        WOOD_PANEL_PARTS_LIST.clear()
        SPECIAL_ORDER_PARTS_LIST.clear()

        F_MATERIAL_PRICES.clear()
        F_HARDWARE_PRICES.clear()
        F_ACCESSORY_PRICES.clear()
        F_WOOD_PANEL_PRICES.clear()
        F_SPECIAL_ORDER_PRICES.clear()
        F_LABOR_PRICES.clear()
        F_MATERIAL_SQUARE_FOOTAGE.clear()
        F_MATERIAL_LINEAR_FOOTAGE.clear()
        F_PROJECT_TOTAL_HARDWARE.clear()
        F_PROJECT_TOTAL_ACCESSORIES.clear()
        F_PROJECT_TOTAL_MATERIAL.clear()
        F_PROJECT_TOTAL_WOOD_PANEL.clear()
        F_PROJECT_TOTAL_SQUARE_FOOTAGE.clear()
        F_PROJECT_TOTAL_LINEAR_FOOTAGE.clear()
        F_PROJECT_TOTAL_PRICE.clear()
        F_ROOM_PRICING_LIST.clear()

    def calculate_price(self, context):
        '''
            This function calculates the price of the project from the xml export
        '''
        xml_file = get_project_xml(self)
        if xml_file:
            calculate_project_price(xml_file)
        props = get_pricing_props()
        if props.export_pricing_parts_list:
            generate_retail_parts_list()
            if bpy.context.preferences.addons['snap'].preferences.enable_franchise_pricing:
                generate_franchise_parts_list()
            

    def draw(self, layout):
        main_box = layout.box()

        row = main_box.row()
        row.scale_y = 1.3
        row.operator(PRICING_PROPERTY_NAMESPACE + ".calculate_price",icon='FILE_TICK')
        main_col = main_box.column(align=True)
        row = main_col.row(align=True)
        row.prop(self,'export_pricing_parts_list',text="Export Pricing Parts List (.xlsx)")
        # row.prop(self,'auto_calculate_on_save',text="Auto Calculate on Save")
                   
        main_col = main_box.column(align=True)
        row = main_col.row(align=True)
        row.scale_y = 1.1
        row.prop_enum(self, "pricing_tabs", 'RETAIL', icon='PREFERENCES', text="Retail Pricing")

        if self.pricing_tabs == 'RETAIL':
            box = main_col.box()
            col = box.column()
            for i in range(len(R_ROOM_PRICING_LIST)):
                col.label(text="Room Name: " + R_ROOM_PRICING_LIST[i][0])
                col.label(text="Special Order Items Count: " + str(R_ROOM_PRICING_LIST[i][7]),icon='BLANK1')
                col.label(text="Material Price: " + sn_unit.draw_dollar_price(R_ROOM_PRICING_LIST[i][3]),icon='BLANK1')
                col.label(text="Hardware Price: " + sn_unit.draw_dollar_price(R_ROOM_PRICING_LIST[i][4]),icon='BLANK1')
                col.label(text="Acccessories Price: " + sn_unit.draw_dollar_price(R_ROOM_PRICING_LIST[i][5]),icon='BLANK1')
                col.label(text="Wood Panel Price: " + sn_unit.draw_dollar_price(R_ROOM_PRICING_LIST[i][6]),icon='BLANK1')
                # col.label(text="Labor Price: " + sn_unit.draw_dollar_price(R_ROOM_PRICING_LIST[i][8]),icon='BLANK1')
                col.label(text="Room Total Price: " + sn_unit.draw_dollar_price(R_ROOM_PRICING_LIST[i][9]),icon='BLANK1')
                col.separator()

            col.separator()
            col.label(text="Project Totals: ")
            col.label(text="Room Count: " + str(len(R_ROOM_PRICING_LIST)),icon='BLANK1')
            col.label(text="Special Order Items Count: " + str(len(SPECIAL_ORDER_PARTS_LIST)),icon='BLANK1')
            col.label(text="Material Price: " + sn_unit.draw_dollar_price(sum(map(float, R_PROJECT_TOTAL_MATERIAL))),icon='BLANK1')
            col.label(text="Hardware Price: " + sn_unit.draw_dollar_price(sum(map(float, R_PROJECT_TOTAL_HARDWARE))),icon='BLANK1')
            col.label(text="Accessories Price: " + sn_unit.draw_dollar_price(sum(map(float, R_PROJECT_TOTAL_ACCESSORIES))),icon='BLANK1')
            col.label(text="Wood Panel Price: " + sn_unit.draw_dollar_price(sum(map(float, R_PROJECT_TOTAL_WOOD_PANEL))),icon='BLANK1')
            # col.label(text="Labor Price: " + sn_unit.draw_dollar_price(sum(map(float, R_PROJECT_TOTAL_LABOR))),icon='BLANK1')
            col.label(text="Project Subtotal: " + sn_unit.draw_dollar_price(sum(map(float, R_PROJECT_TOTAL_PRICE))),icon='BLANK1')

            if len(SPECIAL_ORDER_PARTS_LIST) != 0:
                col.separator()
                col.label(text="This project contains special order items,")
                col.label(text="which are not being calculated in this pricing summary") 
            
        if bpy.context.preferences.addons['snap'].preferences.enable_franchise_pricing:
            row.prop_enum(self, "pricing_tabs", 'FRANCHISE', icon='PREFERENCES', text="Franchise Pricing")

            if self.pricing_tabs == 'FRANCHISE':
                box = main_col.box()
                col = box.column()
                for i in range(len(F_ROOM_PRICING_LIST)):
                    col.label(text="Room Name: " + F_ROOM_PRICING_LIST[i][0])
                    col.label(text="Special Order Items Count: " + str(F_ROOM_PRICING_LIST[i][7]),icon='BLANK1')
                    col.label(text="Material Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][3]),icon='BLANK1')
                    col.label(text="Hardware Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][4]),icon='BLANK1')
                    col.label(text="Acccessories Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][5]),icon='BLANK1')
                    col.label(text="Wood Panel Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][6]),icon='BLANK1')
                    # col.label(text="Labor Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][8]),icon='BLANK1')
                    col.label(text="Room Total Price: " + sn_unit.draw_dollar_price(F_ROOM_PRICING_LIST[i][9]),icon='BLANK1')
                    col.separator()

                col.separator()
                col.label(text="Project Totals: ")
                col.label(text="Room Count: " + str(len(F_ROOM_PRICING_LIST)),icon='BLANK1')
                col.label(text="Special Order Items Count: " + str(len(SPECIAL_ORDER_PARTS_LIST)),icon='BLANK1')
                col.label(text="Material Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_MATERIAL))),icon='BLANK1')
                col.label(text="Hardware Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_HARDWARE))),icon='BLANK1')
                col.label(text="Accessories Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_ACCESSORIES))),icon='BLANK1')
                col.label(text="Wood Panel Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_WOOD_PANEL))),icon='BLANK1')
                # col.label(text="Labor Price: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_LABOR))),icon='BLANK1')
                col.label(text="Project Subtotal: " + sn_unit.draw_dollar_price(sum(map(float, F_PROJECT_TOTAL_PRICE))),icon='BLANK1')

                if len(SPECIAL_ORDER_PARTS_LIST) != 0:
                    col.separator()
                    col.label(text="This project contains special order items,")
                    col.label(text="which are not being calculated in this pricing summary") 
            


class SNAP_PT_Project_Pricing_Setup(Panel):
    bl_label = "Project Pricing"
    bl_space_type = "VIEW_3D"
    bl_region_type = "TOOLS"
    bl_context = "objectmode"
    bl_options = {'DEFAULT_CLOSED'}
    
    def draw_header(self, context):
        layout = self.layout
        layout.label(text='',icon='LINE_DATA')

    def draw(self, context):
        props = get_pricing_props()
        props.draw(self.layout)


# REGISTER CLASSES
def register():
    bpy.utils.register_class(SNAP_OT_Calculate_Price)
    bpy.utils.register_class(SNAP_OT_Select_All_Rooms_Pricing)
    bpy.utils.register_class(SNAP_PROPS_Pricing)
    bpy.utils.register_class(SNAP_PT_Project_Pricing_Setup)
    exec("bpy.types.Scene." + PRICING_PROPERTY_NAMESPACE + "= PointerProperty(type = SNAP_PROPS_Pricing)")


def unregister():
    bpy.utils.unregister_class(SNAP_PT_Project_Pricing_Setup)
    bpy.utils.unregister_class(SNAP_PROPS_Pricing)
    bpy.utils.unregister_class(SNAP_OT_Calculate_Price)
    bpy.utils.unregister_class(SNAP_OT_Select_All_Rooms_Pricing)
    exec("del bpy.types.Scene." + PRICING_PROPERTY_NAMESPACE)


# AUTO CALL OPERATOR ON SAVE
# @bpy.app.handlers.persistent
# def calculate_pricing(scene=None):
#     props = get_pricing_props()
#     if props.auto_calculate_on_save:
#         exec("bpy.ops." + PRICING_PROPERTY_NAMESPACE + ".calculate_price()")
# bpy.app.handlers.save_pre.append(calculate_pricing)